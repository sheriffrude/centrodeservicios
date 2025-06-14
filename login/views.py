import datetime
import json
import uuid
from django.shortcuts import redirect, render
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, connection, connections
import pandas as pd
from django.http import HttpResponse, HttpResponseRedirect

from centrodeservicios import settings

from .forms import UploadFileForm
from django.http import JsonResponse
import openpyxl
from openpyxl import load_workbook
from django.contrib import messages
from django.template.loader import render_to_string
import xlsxwriter
import pdfkit
from django.template import loader
from django.http import FileResponse
import logging
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.cache import never_cache
from uuid import uuid4
import os


from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin

class CustomPasswordChangeView(LoginRequiredMixin, PasswordChangeView):
    template_name = 'registration/password_change_form.html'
    success_url = reverse_lazy('password_change_done')


#---Define La Vista del login-----
def signin(request):
   if request.method == 'GET' :
        return render(request, 'login.html',{
            'form' : AuthenticationForm
            })
   else:
       user = authenticate( username=request.POST['username'], 
                           password=request.POST['password'])
       if user is None or not user.is_active:
            # Usuario no válido o cuenta desactivada
            return render(request, 'login.html', {
                'form': AuthenticationForm,
                'error': 'Usuario o contraseña incorrectos'
    })
       else:
           login (request, user)
           return redirect('home')

#---Define La Vista Principal----
@never_cache
@login_required
def home(request):
    return render(request, 'home.html')





#---Define La Vista del logout-----
@never_cache
@login_required
def exit(request):
    logout(request)
    return redirect('/')
#---Define La Vista del modulo cadena de abastecimiento-----
@never_cache
@login_required
def cadenaabastecimiento(request):
   return render(request, 'cadenaabastecimiento.html')

#---Define La Vistas del modulo Gestion Comercial-----
@never_cache
@login_required
def gestioncomercial(request):
   return render(request, 'gestioncomercial.html')

#---Define La Vistas del modulo Gestion Humana-----

@never_cache
@login_required
def gestionhumana(request):
   return render(request, 'gestionhumana.html')
#---Define La Vistas del modulo Gestion Tecnica-----

@never_cache
@login_required
def gestiontecnica(request):
   return render(request, 'gestiontecnica.html')

@never_cache
@login_required
def despachofrigos(request):
   return render(request, 'despacho_frigos.html')
#---Define La Vistas del modulo Gestion ALIMENTO BALANCEADO-----

@never_cache
@login_required
def gestionalbal(request):
   return render(request, 'gestionalimentobal.html')
#---Define La Vistas del modulo Gestion CALIDAD-----

@never_cache
@login_required
def calidad(request):
   return render(request, 'calidad.html')
#---Define La Vistas del modulo Gestion frigorificos-----

@never_cache
@login_required
def frigorificos(request):
   return render(request, 'frigorificos.html')
#---Define La Vistas del modulo Gestion TI-----

@never_cache
@login_required
def ti(request):
   return render(request, 'tecnologia.html')

#---Define La Vista del modulo Gestion admin y finan-----
@never_cache
@login_required
def adminfinan(request):
   return render(request, 'gestionadminfinan.html')
#---Define La Vista del modulo Gestion admin y finan-----
@never_cache
@login_required
def sig(request):
   return render(request, 'sig.html')

#---Define La Vista del modulo financiera-----
@never_cache
@login_required
def financiera(request):
    grupos = grupos_asociados(request)  
    return render(request, 'financiera.html', {'grupos_asociados': grupos})



#---Define La Vista del modulo reportes-----
@never_cache
@login_required
def repoprove(request):
   return render(request, 'report_prov.html')

@never_cache
@login_required
def carexitosa(request):
   return render(request, 'carga_exitosa.html')

@never_cache
@login_required
def repofina(request):
   return render(request, 'report_finan.html')

#------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------
#---------------CADENA DE ABASTECIMIENTO --------------------------------------------------------------
#------ vista para el cargue de excel en cadena de abastecimiento--------------------------------------
@never_cache
@login_required
def cargar_excel_cadenaabastecimiento(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ca
            with connections['b_ca'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    granja, mes, semana, cantidad_cerdos, año, fecha_corte = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla compromiso_mes
                    cursor.execute(
                        'INSERT INTO compromiso_mes (granja, mes, semana, cantidad_cerdos, año,fecha_corte, GUID, USUARIO) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)',
                        (granja.value, mes.value, semana.value, cantidad_cerdos.value, año.value, fecha_corte.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en compromiso mes exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')

@never_cache
@login_required
def cargar_excel_disponibilidad(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ca
            with connections['b_ca'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    granja, mes, semana, cantidad_cerdos, año = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla disponibilidad_semanal
                    cursor.execute(
                        'INSERT INTO disponibilidad_semanal (granja, mes, semana, cantidad_cerdos, año, GUID, USUARIO) VALUES (%s, %s, %s, %s, %s, %s, %s)',
                        (granja.value, mes.value, semana.value, cantidad_cerdos.value, año.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en Disponibilidad semanal exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en produccion  cerdos  beneficiados--------------------------------------
@never_cache
@login_required
def cargar_excel_cerdosbeneficiados(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ca
            with connections['b_ca'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    CER_BENEF_COLOMBIA,CER_BENEF_EJE_CAFETERO,PARTICIPACION_EJE_CAFETERO,CER_BENEF_CERCAFE,PARTICIPACION_EJE_CAF_CERCAFE,PARTICIPACION_NACIONAL_CERCAFE,FECHA_CORTE = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla PROD_CARNICA_CERDOS_BENEFICIADOS
                    cursor.execute(
                        'INSERT INTO prod_carnica_cerdos_beneficiados (CER_BENEF_COLOMBIA,CER_BENEF_EJE_CAFETERO,PARTICIPACION_EJE_CAFETERO,CER_BENEF_CERCAFE,PARTICIPACION_EJE_CAF_CERCAFE,PARTICIPACION_NACIONAL_CERCAFE,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s)',
                        (CER_BENEF_COLOMBIA.value,CER_BENEF_EJE_CAFETERO.value,PARTICIPACION_EJE_CAFETERO.value,CER_BENEF_CERCAFE.value,PARTICIPACION_EJE_CAF_CERCAFE.value,PARTICIPACION_NACIONAL_CERCAFE.value,FECHA_CORTE.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en PROD_CARNICA_CERDOS_BENEFICIADOS exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en comparatico plantas--------------------------------------
@never_cache
@login_required
def cargar_excel_compaplanta(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ca
            with connections['b_ca'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    PARAMETRO,VALOR,EMPRESA,FECHA_CORTE = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla prod_carnica_comparativo_plantas
                    cursor.execute(
                        'INSERT INTO prod_carnica_comparativo_plantas (PARAMETRO,VALOR,EMPRESA,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s, %s, %s, %s, %s)',
                        (PARAMETRO.value,VALOR.value,EMPRESA.value,FECHA_CORTE.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en prod_carnica_comparativo_plantas exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en COSTO DESPOSTE--------------------------------------
@never_cache
@login_required
def cargar_excel_costodespos(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ca
            with connections['b_ca'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    TIPO_CLIENTE,NUM_CERDOS_DESPOSTADOS,KG_DESPOSTADOS,PESO_PROM_CERDOS,PRECIO_PROM_KG,COSTO_MATERIA_PRIMA,COSTO_MAQUILA,COSTO_KG_MAQUILADO,MAQUILA_SIN_MP,FECHA_CORTE = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla PROD_CARNICA_COSTO_DESPOSTE
                    cursor.execute(
                        'INSERT INTO prod_carnica_costo_desposte (TIPO_CLIENTE,NUM_CERDOS_DESPOSTADOS,KG_DESPOSTADOS,PESO_PROM_CERDOS,PRECIO_PROM_KG,COSTO_MATERIA_PRIMA,COSTO_MAQUILA,COSTO_KG_MAQUILADO,MAQUILA_SIN_MP,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (TIPO_CLIENTE.value,NUM_CERDOS_DESPOSTADOS.value,KG_DESPOSTADOS.value,PESO_PROM_CERDOS.value,PRECIO_PROM_KG.value,COSTO_MATERIA_PRIMA.value,COSTO_MAQUILA.value,COSTO_KG_MAQUILADO.value,MAQUILA_SIN_MP.value,FECHA_CORTE.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en PROD_CARNICA_COSTO_DESPOSTE exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel enKG_BENEFICIO--------------------------------------
@never_cache
@login_required
def cargar_excel_kgbeneficio(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ca
            with connections['b_ca'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    CER_BENEF_COLOMBIA,CER_BENEF_EJE_CAFETERO,PARTICIPACION_EJE_CAFETERO,CER_BENEF_CERCAFE,PARTICIPACION_EJE_CAF_CERCAFE,PARTICIPACION_NACIONAL_CERCAFE,PESO_CF_NACIONAL,PESO_EJE_CAFETERO,PESO_CF_CERCAFE,KG_NACIONAL,KG_EJE_CAFETERO,KG_CERCAFE,FECHA_CORTE = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla PROD_CARNICA_KG_BENEFICIO
                    cursor.execute(
                        'INSERT INTO prod_carnica_kg_beneficio (CER_BENEF_COLOMBIA,CER_BENEF_EJE_CAFETERO,PARTICIPACION_EJE_CAFETERO,CER_BENEF_CERCAFE,PARTICIPACION_EJE_CAF_CERCAFE,PARTICIPACION_NACIONAL_CERCAFE,PESO_CF_NACIONAL,PESO_EJE_CAFETERO,PESO_CF_CERCAFE,KG_NACIONAL,KG_EJE_CAFETERO,KG_CERCAFE,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (CER_BENEF_COLOMBIA.value,CER_BENEF_EJE_CAFETERO.value,PARTICIPACION_EJE_CAFETERO.value,CER_BENEF_CERCAFE.value,PARTICIPACION_EJE_CAF_CERCAFE.value,
                          PARTICIPACION_NACIONAL_CERCAFE.value,PESO_CF_NACIONAL.value,PESO_EJE_CAFETERO.value,PESO_CF_CERCAFE.value,KG_NACIONAL.value,KG_EJE_CAFETERO.value,KG_CERCAFE.value,FECHA_CORTE.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en PROD_CARNICA_KG_BENEFICIO exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en KG_DESPOSTADOS--------------------------------------
@never_cache 
@login_required
def cargar_excel_kgdesposte(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ca
            with connections['b_ca'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    KG_PRODUCIDOS_CERCAFE,KG_DESPOSTADOS_CERCAFE,PORCENTAJE_PARTICIPACION,TRIMESTRE_2022_CERCAFE,TRIMESTRE_2022_DESPOSTE,TRIMESTRE_2023_CERCAFE,TRIMESTRE_2023_DESPOSTE,CERCIMIENTO_22_23,FECHA_CORTE = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla PROD_CARNICA_KG_DESPOSTADOS
                    cursor.execute(
                        'INSERT INTO prod_carnica_kg_despostados (KG_PRODUCIDOS_CERCAFE,KG_DESPOSTADOS_CERCAFE,PORCENTAJE_PARTICIPACION,TRIMESTRE_2022_CERCAFE,TRIMESTRE_2022_DESPOSTE,TRIMESTRE_2023_CERCAFE,TRIMESTRE_2023_DESPOSTE,CERCIMIENTO_22_23,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (KG_PRODUCIDOS_CERCAFE.value,KG_DESPOSTADOS_CERCAFE.value,PORCENTAJE_PARTICIPACION.value,TRIMESTRE_2022_CERCAFE.value,TRIMESTRE_2022_DESPOSTE.value,TRIMESTRE_2023_CERCAFE.value,
                          TRIMESTRE_2023_DESPOSTE.value,CERCIMIENTO_22_23.value,FECHA_CORTE.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en PROD_CARNICA_KG_DESPOSTADOS exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en PARTICIPACION_CORTES--------------------------------------
@never_cache 
@login_required
def cargar_excel_particortes(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ca
            with connections['b_ca'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    CORTE,PORCENTAJE_PARTICIPACION,PORCENTAJE_META,PESO_PROM_CANAL,CANTIDAD_CANALES,FECHA_CORTE = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla PROD_CARNICA_PARTICIPACION_CORTES
                    cursor.execute(
                        'INSERT INTO prod_carnica_participacion_cortes (CORTE,PORCENTAJE_PARTICIPACION,PORCENTAJE_META,PESO_PROM_CANAL,CANTIDAD_CANALES,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s)',
                        (CORTE.value,PORCENTAJE_PARTICIPACION.value,PORCENTAJE_META.value,PESO_PROM_CANAL.value,CANTIDAD_CANALES.value,FECHA_CORTE.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en PROD_CARNICA_PARTICIPACION_CORTES exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en TON_IMPORTADAS--------------------------------------
@never_cache 
@login_required
def cargar_excel_toneladasimport(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ca
            with connections['b_ca'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    CER_BENEF_COLOMBIA,TON_BENEF_COLOMBIA,TON_IMPORT_COLOMBIA,CERDOS_IMPORTADOS,ENE_FEB_22_TON_BENEF,ENE_FEB_23_TON_BENEF,CRECIMIENTO_22_23,ENE_FEB_MAR_22_TON_IMPORT,ENE_FEB_MAR_23_TON_IMPORT,CRECIMIENTO_OMPORT_22_23,FECHA_CORTE = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla PROD_CARNICA_TON_IMPORTADAS
                    cursor.execute(
                        'INSERT INTO prod_carnica_ton_importadas (CER_BENEF_COLOMBIA,TON_BENEF_COLOMBIA,TON_IMPORT_COLOMBIA,CERDOS_IMPORTADOS,ENE_FEB_22_TON_BENEF,ENE_FEB_23_TON_BENEF,CRECIMIENTO_22_23,ENE_FEB_MAR_22_TON_IMPORT,ENE_FEB_MAR_23_TON_IMPORT,CRECIMIENTO_OMPORT_22_23,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (CER_BENEF_COLOMBIA.value,TON_BENEF_COLOMBIA.value,TON_IMPORT_COLOMBIA.value,CERDOS_IMPORTADOS.value,ENE_FEB_22_TON_BENEF.value,ENE_FEB_23_TON_BENEF.value,CRECIMIENTO_22_23.value,ENE_FEB_MAR_22_TON_IMPORT.value,ENE_FEB_MAR_23_TON_IMPORT.value,CRECIMIENTO_OMPORT_22_23.value,FECHA_CORTE.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en PROD_CARNICA_TON_IMPORTADAS exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')






#------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------
#------------------ CARGA DE GESTION COMERCIAL --------------------------------------------------------
#------------------------------------------------------------------------------------------------------
#------ vista para el cargue de excel en clientes activos----------------------------------------------
@never_cache
@login_required
def cargar_excel_clientesactivos(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gc
            with connections['b_gc'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    FECHA_CORTE, CANTIDAD_CLIENTES, ZONA_CLIENTE, KG_FACTURADOS,DINERO_APORTADO,ESTADO_CLIENTE = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla compromiso_mes
                    cursor.execute(
                        'INSERT INTO clientes_activos (FECHA_CORTE,CANTIDAD_CLIENTES,ZONA_CLIENTE,KG_FACTURADOS,DINERO_APORTADO,ESTADO_CLIENTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s,%s,%s)',
                        (FECHA_CORTE.value, CANTIDAD_CLIENTES.value, ZONA_CLIENTE.value, KG_FACTURADOS.value,DINERO_APORTADO.value,ESTADO_CLIENTE.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en CLIENTES ACTIVOS exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en clientes----------------------------------------------
# @never_cache
# @login_required
# def cargar_excel_clientesactivos(request):
#     if request.method == 'POST':
#         try:
#             archivo_excel = request.FILES['archivo_excel']
#             wb = openpyxl.load_workbook(archivo_excel)
#             ws = wb.active
#             guid = str(uuid4())
#             usuario = request.user
#             # Abre una conexión a la base de datos b_gc
#             with connections['dhc'].cursor() as cursor:
#                 for row in ws.iter_rows(min_row=2):
#                     NIT,RAZON_SOCIAL,CUPO,DIRECCION_SEDE_PRINCIPAL,DIRECCION_EXPENDIO,ID_CLASIFICACION,ID_MUNICIPIO,ID_DEPARTAMENTO,ID_REGION, ID_VENDEDOR,ID_SEGMENTO,ID_linea_negocio = row
#                     # Ejecuta una consulta SQL para insertar los datos en la tabla compromiso_mes
#                     cursor.execute(
#                         'INSERT INTO clientes (NIT,RAZON_SOCIAL,CUPO,DIRECCION_SEDE_PRINCIPAL,DIRECCION_EXPENDIO,ID_CLASIFICACION,ID_MUNICIPIO,ID_DEPARTAMENTO,ID_REGION, ID_VENDEDOR,ID_SEGMENTO,ID_linea_negocio,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s,%s,%s,%s,%s,%s,%s,%s,%s)',
#                         (NIT.value,RAZON_SOCIAL.value,CUPO.value,DIRECCION_SEDE_PRINCIPAL.value,DIRECCION_EXPENDIO.value,ID_CLASIFICACION.value,ID_MUNICIPIO.value,ID_DEPARTAMENTO.value,ID_REGION.value, ID_VENDEDOR.value,ID_SEGMENTO.value,ID_linea_negocio.value,guid,usuario.username)
#                     )
#             messages.success(request, 'Carga de datos en CLIENTES  exitosa')
#         except KeyError:
#             messages.error(request, 'No se ha proporcionado un archivo Excel.')
#         except IntegrityError as e:
#             messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
#         except Exception as e:
#             messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
#         return redirect('home')
#     return render(request, '/home/')
#------ vista para el cargue de excel en ventas---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_ventas(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gc
            with connections['b_gc'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    FECHA_CORTE,LINEA_NEGOCIO,PRESUPUESTO_UNIDADES,PRESUPUESTO_KG,UNIDADES_VENDIDAS,KG_VENDIDO,VALOR_VENTA,PRESUPUESTO_VENTA= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla VENTAS
                    cursor.execute(
                        'INSERT INTO ventas (FECHA_CORTE,LINEA_NEGOCIO,PRESUPUESTO_UNIDADES,PRESUPUESTO_KG,UNIDADES_VENDIDAS,KG_VENDIDO,VALOR_VENTA,PRESUPUESTO_VENTA,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s,%s,%s,%s,%s)',
                        (FECHA_CORTE.value, LINEA_NEGOCIO.value, PRESUPUESTO_UNIDADES.value, PRESUPUESTO_KG.value,UNIDADES_VENDIDAS.value,KG_VENDIDO.value,VALOR_VENTA.value,PRESUPUESTO_VENTA.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en VENTAS exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
# ----------------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#--------------------------------- CARGA DE GESTION HUMANA -------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#------ vista para el cargue de excel en NOMINA---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_nomina(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gh
            with connections['b_gh'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    FECHA_CORTE,AREA,CENTRO_COSTO,NUM_COLABORADORES,COSTO_PROV= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla VENTAS
                    cursor.execute(
                        'INSERT INTO nomina (FECHA_CORTE,AREA,CENTRO_COSTO,NUM_COLABORADORES,COSTO_PROV,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s,%s)',
                        (FECHA_CORTE.value, AREA.value, CENTRO_COSTO.value, NUM_COLABORADORES.value,COSTO_PROV.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en NOMINA exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: archivo diferente a la plantilla')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en PROMOCIONES---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_promo(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gh
            with connections['b_gh'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    FECHA_CORTE,NOMBRE,ANTIGUO_CARGO,NUEVO_CARGO= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla VENTAS
                    cursor.execute(
                        'INSERT INTO promociones (FECHA_CORTE,NOMBRE,ANTIGUO_CARGO,NUEVO_CARGO,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s)',
                        (FECHA_CORTE.value, NOMBRE.value, ANTIGUO_CARGO.value, NUEVO_CARGO.value,guid,usuario.usuario)
                    )
            messages.success(request, 'Carga de datos en PROMOCIONES exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en PROCESOS DE SELECCION---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_prosele(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gh
            with connections['b_gh'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    
                    NUM_REQUISICION,FECHA_APROBACION,AREA_CENTRO_COSTO,FECHA_RETIRO,NOMBRE_RETIRADO,CARGO,CUBRIMIENTO_ESPERADO_DIAS,NOMBRE_CANDIDATO,TIPO_INGRESO_PROMO_INT,EXAMEN_MEDICO,VISITA_DOMICILIARIA,POLIGRAFIA,FECHA_INGRESO = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla PROCESO_SELECCION
                    cursor.execute(
                        'INSERT INTO proceso_seleccion (NUM_REQUISICION,FECHA_APROBACION,AREA_CENTRO_COSTO,FECHA_RETIRO,NOMBRE_RETIRADO,CARGO,CUBRIMIENTO_ESPERADO_DIAS,NOMBRE_CANDIDATO,TIPO_INGRESO_PROMO_INT,EXAMEN_MEDICO,VISITA_DOMICILIARIA,POLIGRAFIA,FECHA_INGRESO,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (NUM_REQUISICION.value,FECHA_APROBACION.value,AREA_CENTRO_COSTO.value,FECHA_RETIRO.value,NOMBRE_RETIRADO.value,CARGO.value,CUBRIMIENTO_ESPERADO_DIAS.value,NOMBRE_CANDIDATO.value,TIPO_INGRESO_PROMO_INT.value,EXAMEN_MEDICO.value,VISITA_DOMICILIARIA.value,POLIGRAFIA.value,FECHA_INGRESO.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en PROCESO SELECCION exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en RETENCION ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_retencion(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gh
            with connections['b_gh'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    FECHA_REPORTE,INDICADOR_RETENCION,OBSERVACIONES= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla RETENCION
                    cursor.execute(
                        'INSERT INTO retencion (FECHA_REPORTE,INDICADOR_RETENCION,OBSERVACIONES,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s)',
                        (FECHA_REPORTE.value,INDICADOR_RETENCION.value,OBSERVACIONES.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en RETENCION exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en ROTACION ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_rotacion(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gh
            with connections['b_gh'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    FECHA_REPORTE,INDICADOR_ROTACION,OBSERVACIONES= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla VENTAS
                    cursor.execute(
                        'INSERT INTO rotacion (FECHA_REPORTE,INDICADOR_ROTACION,OBSERVACIONES,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s)',
                        (FECHA_REPORTE.value, INDICADOR_ROTACION.value, OBSERVACIONES.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en ROTACION exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en SST DIAGNOSTICOS INDICADORES ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_sstdiag(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gh
            with connections['b_gh'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    FECHA_CORTE,SEDE,DIAGNOSTICO,CANTIDAD,OBSERVACION= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla SST_DIAGNOSTICOS_INDICADORES
                    cursor.execute(
                        'INSERT INTO sst_diagnosticos_indicadores (FECHA_CORTE,SEDE,DIAGNOSTICO,CANTIDAD,OBSERVACION,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s)',
                        (FECHA_CORTE.value, SEDE.value, DIAGNOSTICO.value,CANTIDAD.value,OBSERVACION.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en SST DIAGNOSTICOS INDICADORES exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')

#------ vista para el cargue de excel en Recuperado nomina ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_recunomina(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gh
            with connections['b_gh'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    VALOR_RECUPERADO,FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla Recuperado nomina
                    cursor.execute(
                        'INSERT INTO recuperado_nomina (VALOR_RECUPERADO,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s)',
                        (VALOR_RECUPERADO.value, FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en Recuperado nomina exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')


#------ vista para el cargue de excel en SST INDICADORES ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_sstindi(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gh
            with connections['b_gh'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                   
                    FECHA_CORTE,SEDE,CANTIDAD_PEG,DIAS_INCAPACIDAD_PEL,CANTIDAD_PAT,PRORROGAS,DIAS_INCAPACIDAD_PAT,LICENCIA_MATERNIDAD,DIAS_LICENCIA_MAT,LICENCIA_PATERNIDAD,DIAS_LICENCIA_PAT,COSTO_INCAPACIDAD,OBSERVACIONES= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla VENTAS
                    cursor.execute(
                        'INSERT INTO sst_indicadores (FECHA_CORTE,SEDE,CANTIDAD_PEG,DIAS_INCAPACIDAD_PEL,CANTIDAD_PAT,PRORROGAS,DIAS_INCAPACIDAD_PAT,LICENCIA_MATERNIDAD,DIAS_LICENCIA_MAT,LICENCIA_PATERNIDAD,DIAS_LICENCIA_PAT,COSTO_INCAPACIDAD,OBSERVACIONES,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s)',
                        (FECHA_CORTE.value,SEDE.value,CANTIDAD_PEG.value,DIAS_INCAPACIDAD_PEL.value,CANTIDAD_PAT.value,PRORROGAS.value,DIAS_INCAPACIDAD_PAT.value,LICENCIA_MATERNIDAD.value,DIAS_LICENCIA_MAT.value,LICENCIA_PATERNIDAD.value,DIAS_LICENCIA_PAT.value,COSTO_INCAPACIDAD.value,OBSERVACIONES.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en SST INDICADORES exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en SST SEVERIDAD Y FRECUENCIA ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_sstseveridad(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gh
            with connections['b_gh'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                   
                    FECHA_CORTE,CANT_ENF_GENERAL,CANT_ACC_TRABAJO,NUM_EMPLEADOS,FREC_ACC,DIAS_INC_GENERAL,DIAS_INC_ACC,SEV_ACC,INCID_ENF_LAB,PORC_AUSENTISMO= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla SST_SEVERIDAD_Y_FRECUENCIA
                    cursor.execute(
                        'INSERT INTO sst_severidad_y_frecuencia (FECHA_CORTE,CANT_ENF_GENERAL,CANT_ACC_TRABAJO,NUM_EMPLEADOS,FREC_ACC,DIAS_INC_GENERAL,DIAS_INC_ACC,SEV_ACC,INCID_ENF_LAB,PORC_AUSENTISMO,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s,%s, %s, %s, %s, %s)',
                        (FECHA_CORTE.value,CANT_ENF_GENERAL.value,CANT_ACC_TRABAJO.value,NUM_EMPLEADOS.value,FREC_ACC.value,DIAS_INC_GENERAL.value,DIAS_INC_ACC.value,SEV_ACC.value,INCID_ENF_LAB.value,PORC_AUSENTISMO.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en SST SEVERIDAD Y FRECUENCIA exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
# ----------------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#--------------------------------- CARGA DE GESTION TECNICA -------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#------ vista para el cargue de excel en Abastecimiento Hembras ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_abashem(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gt
            with connections['b_gt'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                   
                    GRANJA,CANTIDAD_ENTREGADA,PORCENTAJE_CUMPLIMIENTO,FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla SST_SEVERIDAD_Y_FRECUENCIA
                    cursor.execute(
                        'INSERT INTO abastecimiento_hembras (GRANJA,CANTIDAD_ENTREGADA,PORCENTAJE_CUMPLIMIENTO,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s)',
                        (GRANJA.value,CANTIDAD_ENTREGADA.value,PORCENTAJE_CUMPLIMIENTO.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en ABASTECIMIENTO HEMBRAS exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')

#------ vista para el cargue de excel en FORTUITOS ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_fortuitos(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gt
            with connections['b_gt'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                  
                    FECHA_CORTE,PLANTA,GRANJA,CANTIDAD_MUERTE_TRANSPORTE,CANTIDAD_MUERTE_REPOSO,AGITADOS,LESIONADOS,CAIDOS,RETOMAS,TOTAL= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla SST_SEVERIDAD_Y_FRECUENCIA
                    cursor.execute(
                        'INSERT INTO fortuitos3 (FECHA_CORTE,PLANTA,GRANJA,CANTIDAD_MUERTE_TRANSPORTE,CANTIDAD_MUERTE_REPOSO,AGITADOS,LESIONADOS,CAIDOS,RETOMAS,TOTAL,GUID,USUARIO) VALUES (%s, %s,%s, %s, %0s, %s, %s,%s, %s, %s,%s,%s)',
                        (FECHA_CORTE.value,PLANTA.value,GRANJA.value,CANTIDAD_MUERTE_TRANSPORTE.value,CANTIDAD_MUERTE_REPOSO.value,AGITADOS.value,LESIONADOS.value,CAIDOS.value,RETOMAS.value,TOTAL.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en FORTUITOS exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en KG VENDIDOS HEMBRA ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_kgvend(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gt
            with connections['b_gt'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                  
                    GRANJA,KG_V_H_A,ASOCIADO,FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla kg_vendidos_hembra
                    cursor.execute(
                        'INSERT INTO kg_vendidos_hembra (GRANJA,KG_V_H_A,ASOCIADO,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s)',
                        (GRANJA.value,KG_V_H_A.value,ASOCIADO.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en KG VENDIDOS HEMBRA exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en PESO FINAL CONVERSION ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_pesofinconver(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user

            # Abre una conexión a la base de datos b_gt
            with connections['b_gt'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                  
                    GRANJA,PESO,META_PESO,CONVERSION_META,CONVERSION,FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla peso_final_conversion
                    cursor.execute(
                        'INSERT INTO peso_final_conversion (GRANJA,PESO,META_PESO,CONVERSION_META,CONVERSION,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s)',
                        (GRANJA.value,PESO.value,META_PESO.value,CONVERSION_META.value,CONVERSION.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en PESO FINAL CONVERSION exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en PROYECCION HEMBRAS ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_proyhem(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gt
            with connections['b_gt'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    
                    PARTOS,TASA_PARTOS,CUMPLIMIENTO_PROYECTADO,CUMPLIMIENTO_REAL,AÑO_SERVICIO,OBSERVACIONES,FECHA_CORTE,CANTIDAD_PROYECCION_HEMBRAS,CANTIDAD_REAL= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla PROYECCION_HEMBRAS
                    cursor.execute(
                        'INSERT INTO proyeccion_hembras (PARTOS,TASA_PARTOS,CUMPLIMIENTO_PROYECTADO,CUMPLIMIENTO_REAL,AÑO_SERVICIO,OBSERVACIONES,FECHA_CORTE,CANTIDAD_PROYECCION_HEMBRAS,CANTIDAD_REAL,GUID,USUARIO) VALUES (%s, %s,%s, %s,%s, %s, %s, %s, %s, %s, %s)',
                        (PARTOS.value,TASA_PARTOS.value,CUMPLIMIENTO_PROYECTADO.value,CUMPLIMIENTO_REAL.value,AÑO_SERVICIO.value,OBSERVACIONES.value,FECHA_CORTE.value,CANTIDAD_PROYECCION_HEMBRAS.value,CANTIDAD_REAL.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en PROYECCION HEMBRAS exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en TECNICA CIA ---------------------------------------------------------------
@never_cache
@login_required
def cargar_excel_tecnicacia(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user

            # Abre una conexión a la base de datos b_gt
            with connections['b_gt'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    
                    LINEA_GENETICA,CANTIDAD_MACHOS,PORCENTAJE_DISTRIBUCION_MACHOS,CANTIDAD_DESECHADO,PORCENTAJE_DESCECHADO,DOSIS_PRODUCIDAS,DOSIS_VENDIDAS,PROMEDIO_MORFOLOGIA,OBSERVACION,FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla TECNICA_CIA
                    cursor.execute(
                        'INSERT INTO tecnica_cia (LINEA_GENETICA,CANTIDAD_MACHOS,PORCENTAJE_DISTRIBUCION_MACHOS,CANTIDAD_DESECHADO,PORCENTAJE_DESCECHADO,DOSIS_PRODUCIDAS,DOSIS_VENDIDAS,PROMEDIO_MORFOLOGIA,OBSERVACION,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (LINEA_GENETICA.value,CANTIDAD_MACHOS.value,PORCENTAJE_DISTRIBUCION_MACHOS.value,CANTIDAD_DESECHADO.value,PORCENTAJE_DESCECHADO.value,DOSIS_PRODUCIDAS.value,DOSIS_VENDIDAS.value,PROMEDIO_MORFOLOGIA.value,OBSERVACION.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en TECNICA CIA exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
# ----------------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#--------------------------------- CARGA DE GESTION ALIMENTO BALANCEADO -----------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#-------- vista para el cargue de excel en Planta Alimento Balanceado -----------------------------------------------
@never_cache
@login_required
def cargar_excel_alibal(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gt
            with connections['b_gab'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    
                    TONELADAS_PRODUCIDAS_MES,TONELADAS_PRESUPUESTO_MES,PORCENTAJE_VARIACION_MES,PORCENTAJE_CUMPLIMIENTO_MES,OBSERVACION_VARIACION,PORCENTAJE_BULTO_MES,PORCENTAJE_GRANEL_MES,SACK_OFF,PORCENTAJE_OTIF,OBSERVACION_OTIF,PRESUPUESTO_MO_CIF,MO_CIF,TIEMPO_MUERTO,COSTO_TIEMPO_MUERTO,OBSERVACION_TIEMPO_MUERTO,FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla PLANTA_ALIMENTOS_BALANCEADOS
                    cursor.execute(
                        'INSERT INTO planta_alimentos_balanceados (TONELADAS_PRODUCIDAS_MES,TONELADAS_PRESUPUESTO_MES,PORCENTAJE_VARIACION_MES,PORCENTAJE_CUMPLIMIENTO_MES,OBSERVACION_VARIACION,PORCENTAJE_BULTO_MES,PORCENTAJE_GRANEL_MES,SACK_OFF,PORCENTAJE_OTIF,OBSERVACION_OTIF,PRESUPUESTO_MO_CIF,MO_CIF,TIEMPO_MUERTO,COSTO_TIEMPO_MUERTO,OBSERVACION_TIEMPO_MUERTO,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (TONELADAS_PRODUCIDAS_MES.value,TONELADAS_PRESUPUESTO_MES.value,PORCENTAJE_VARIACION_MES.value,PORCENTAJE_CUMPLIMIENTO_MES.value,OBSERVACION_VARIACION.value,PORCENTAJE_BULTO_MES.value,PORCENTAJE_GRANEL_MES.value,SACK_OFF.value,PORCENTAJE_OTIF.value,OBSERVACION_OTIF.value,PRESUPUESTO_MO_CIF.value,MO_CIF.value,TIEMPO_MUERTO.value,COSTO_TIEMPO_MUERTO.value,OBSERVACION_TIEMPO_MUERTO.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en PLANTA ALIMENTOS BALANCEADOS exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------
#------------------ CARGA DE BSC ----------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------
#------ vista para el cargue de excel en BSC-----------------------------------------------------------
@never_cache
@login_required
def cargar_excel_bsc(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gc
            with connections['b_sig'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    PERSPECTIVA, ESTRATEGIA, INDICADOR, META, META_NUMERICA, MES, UNIDAD_MEDIDA, RESULTADO, RESULTADO_NUMERICO, RESULTADO_GENERAL = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla bsc
                    cursor.execute(
                        'INSERT INTO bsc (PERSPECTIVA,ESTRATEGIA,INDICADOR,META,META_NUMERICA,MES,UNIDAD_MEDIDA,RESULTADO,RESULTADO_NUMERICO,RESULTADO_GENERAL,GUID,USUARIO) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (PERSPECTIVA.value,ESTRATEGIA.value,INDICADOR.value,META.value,META_NUMERICA.value,MES.value,UNIDAD_MEDIDA.value,RESULTADO.value,RESULTADO_NUMERICO.value,RESULTADO_GENERAL.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en bsc exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
# ----------------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#--------------------------------- CARGA DE   CALIDAD ------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#-------- vista para el cargue de excel en Avance Proceso --------------------------------------------------------
@never_cache
@login_required
def cargar_excel_avancepro(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_c
            with connections['b_c'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    TIPO,PROCESO,DETALLE_PROCESO,AVANCE,META,FECHA_CORTE,_= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla avance_proceso
                    cursor.execute(
                        'INSERT INTO avance_proceso (TIPO,PROCESO,DETALLE_PROCESO,AVANCE,META,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s)',
                        (TIPO.value,PROCESO.value,DETALLE_PROCESO.value,AVANCE.value,META.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en AVANCE PROCESO exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#-------- vista para el cargue de excel en Calidad Planta --------------------------------------------------------
@never_cache
@login_required
def cargar_excel_calidadpl(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user

            # Abre una conexión a la base de datos b_c
            with connections['b_c'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    PORCENTAJE_DESVIACIONES_CALIDAD,TONELADAS_REPROCESADAS,TONELADAS_LIBERADAS_CONCESION,PORCENTAJE_RETENCION,PORCENTAJE_MEZCLA,PORCENTAJE_DURABILIDAD,PORCENTAJE_FINOS,PORCENTAJE_FORMULACION,CUMPLIMIENTO_BPM,FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla CALIDAD_PLANTA
                    cursor.execute(
                        'INSERT INTO calidad_planta (PORCENTAJE_DESVIACIONES_CALIDAD,TONELADAS_REPROCESADAS,TONELADAS_LIBERADAS_CONCESION,PORCENTAJE_RETENCION,PORCENTAJE_MEZCLA,PORCENTAJE_DURABILIDAD,PORCENTAJE_FINOS,PORCENTAJE_FORMULACION,CUMPLIMIENTO_BPM,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (PORCENTAJE_DESVIACIONES_CALIDAD.value,TONELADAS_REPROCESADAS.value,TONELADAS_LIBERADAS_CONCESION.value,PORCENTAJE_RETENCION.value,PORCENTAJE_MEZCLA.value,PORCENTAJE_DURABILIDAD.value,PORCENTAJE_FINOS.value,PORCENTAJE_FORMULACION.value,CUMPLIMIENTO_BPM.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en CALIDAD PLANTA exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')

#-------- vista para el cargue de excel en CAUSAS DESVIACIONES --------------------------------------------------------
@never_cache
@login_required
def cargar_excel_causasdes(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_c
            with connections['b_c'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    CAUSA,PLAN_ACCION,TON_REPROCESADAS,FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla CAUSAS_DESVIACIONES
                    cursor.execute(
                        'INSERT INTO causas_desviaciones (CAUSA,PLAN_ACCION,TON_REPROCESADAS,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s)',
                        (CAUSA.value,PLAN_ACCION.value,TON_REPROCESADAS.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en CAUSAS_DESVIACIONES exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')






#-------- vista para el cargue de excel en oinc --------------------------------------------------------

@never_cache
@login_required
def cargar_excel_ingresoinc(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_c
            with connections['oinc'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    mes,año,semana,lote,lote_cod_canal,f_ingreso,solicitante,propietario,granja,rem_granja,rem_solicitante,mun_granja,guia_ica,verificacion_ica, laboratorio_ic, registro_ic,cod_canal,genero,prom_peso_pie,observacion= row

                    # Ejecuta una consulta SQL para insertar los datos en la tabla CAUSAS_DESVIACIONES
                    cursor.execute(
                        'INSERT INTO ingreso (mes,año,semana,lote,lote_cod_canal,f_ingreso,solicitante,propietario,granja,rem_granja,rem_solicitante,mun_granja,guia_ica,verificacion_ica, laboratorio_ic, registro_ic,cod_canal,genero,prom_peso_pie,observacion,GUID,USUARIO) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                        (mes.value ,año.value ,semana.value ,lote.value ,lote_cod_canal.value ,f_ingreso.value ,solicitante.value ,propietario.value ,granja.value ,rem_granja.value ,rem_solicitante.value ,mun_granja.value ,guia_ica.value ,verificacion_ica.value , laboratorio_ic.value , registro_ic.value ,cod_canal.value ,genero.value ,prom_peso_pie.value ,observacion.value ,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en ingreso exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')


@never_cache
@login_required
def cargar_excel_despachoinc(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_c
            with connections['oinc'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    mes,año,semana,lote,lote_cod_canal,fase,f_ingreso,solicitante, propietario,granja,rem_solicitante,c_fria,clas_abc,f_remision,destino_cliente,destino_remision, direccion_remision, municipio_destino,departamento_destino,remision,tipo_remision,g_invima,placa_furgon,observacion= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla CAUSAS_DESVIACIONES
                    cursor.execute(
                        'INSERT INTO despacho (mes,año,semana,lote,lote_cod_canal,fase,f_ingreso,solicitante, propietario,granja,rem_solicitante,c_fria,clas_abc,f_remision,destino_cliente,destino_remision, direccion_remision, municipio_destino,departamento_destino,remision,tipo_remision,g_invima,placa_furgon,observacion,GUID,USUARIO) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                        (mes.value,año.value,semana.value,lote.value,lote_cod_canal.value,fase.value,f_ingreso.value,solicitante.value, propietario.value,granja.value,rem_solicitante.value,c_fria.value,clas_abc.value,f_remision.value,destino_cliente.value,destino_remision.value, direccion_remision.value, municipio_destino.value,departamento_destino.value,remision.value,tipo_remision.value,g_invima.value,placa_furgon.value,observacion.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en despacho exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')


@never_cache
@login_required
def cargar_excel_beneficiorendimientoinc(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_c
            with connections['oinc'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    mes,año, semana, lote, lote_turno_beneficio, lote_cod_canal, fase, f_ingreso, f_beneficio, f_vencimiento, solicitante, propietario, granja, rem_solicitante, mun_granja, guia_ica, verificacion_ica, laboratorio_ic, registro_ic, turno_beneficio, cod_canal, presentacion, c_caliente, grasa_dorsal, rto_pcc, c_fria, rto_pcf, clasificacion, merma, magro, cava, tiempo_cava, clas_seurop, clase_abc, factura_beneficio, observacion= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla CAUSAS_DESVIACIONES
                    cursor.execute(
                        'INSERT INTO beneficio_rendimiento (mes,año, semana, lote, lote_turno_beneficio, lote_cod_canal, fase, f_ingreso, f_beneficio, f_vencimiento, solicitante, propietario, granja, rem_solicitante, mun_granja, guia_ica, verificacion_ica, laboratorio_ic, registro_ic, turno_beneficio, cod_canal, presentacion, c_caliente, grasa_dorsal, rto_pcc, c_fria, rto_pcf, clasificacion, merma, magro, cava, tiempo_cava, clas_seurop, clase_abc, factura_beneficio, observacion,GUID,USUARIO) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                        (mes.value,año.value, semana.value, lote.value, lote_turno_beneficio.value, lote_cod_canal.value, fase.value, f_ingreso.value, f_beneficio.value, f_vencimiento.value, solicitante.value, propietario.value, granja.value, rem_solicitante.value, mun_granja.value, guia_ica.value, verificacion_ica.value, laboratorio_ic.value, registro_ic.value, turno_beneficio.value, cod_canal.value, presentacion.value, c_caliente.value, grasa_dorsal.value, rto_pcc.value, c_fria.value, rto_pcf.value, clasificacion.value, merma.value, magro.value, cava.value, tiempo_cava.value, clas_seurop.value, clase_abc.value, factura_beneficio.value, observacion.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en beneficio rendimiento exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')




@never_cache
@login_required
def cargar_excel_decomisoinc(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_c
            with connections['oinc'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    lote,f_ingreso,f_beneficio,solicitante,propietario,granja,lote_turno_beneficio,lote_cod_canal, turno_beneficio, cod_canal, acta_decomiso, f_decomiso, decomiso, patologias,cantidad,unidad_m = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla decomisos
                    cursor.execute(
                        'INSERT INTO decomisos (lote,f_ingreso,f_beneficio,solicitante,propietario,granja,lote_turno_beneficio,lote_cod_canal, turno_beneficio, cod_canal, acta_decomiso, f_decomiso, decomiso, patologias,cantidad,unidad_m,GUID,USUARIO) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                        (lote.value, f_ingreso.value, f_beneficio.value, solicitante.value, propietario.value, granja.value, lote_turno_beneficio.value, lote_cod_canal.value, turno_beneficio.value, cod_canal.value, acta_decomiso.value, f_decomiso.value, decomiso.value, patologias.value, cantidad.value, unidad_m.value, guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en decomisos exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
























from datetime import datetime as dt, timedelta
from openpyxl.cell.cell import Cell

@never_cache
@login_required
def cargar_excel_oinc(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos oinc
            with connections['oinc'].cursor() as cursor:
                for row in ws.iter_rows(min_row=4):
                    print(row)
                    # Obtener solo las primeras 52 columnas de la fila
                    row_data = row[:52]
                    # Acceder a los valores de las celdas
                    Mes = row_data[0].value
                    Año = row_data[1].value
                    Semana = row_data[2].value
                    Lote = row_data[3].value
                    Lote_Turn_Bene = row_data[4].value
                    Lote_Cod_Canal = row_data[5].value
                    Fase = row_data[6].value
                    F_Ingreso = row_data[7].value
                    F_Beneficio = row_data[8].value
                    F_Vencimiento = row_data[9].value
                    Solicitante = row_data[10].value
                    Propietario = row_data[11].value
                    Granja = row_data[12].value
                    Rem_Granja = row_data[13].value
                    Rem_Solicitante = row_data[14].value
                    Mun_Granja = row_data[15].value
                    Guia_ICA = row_data[16].value
                    Verificacion_ICA = row_data[17].value
                    Lab_IC = row_data[18].value
                    Registro_IC = row_data[19].value
                    Turno_Beneficio = row_data[20].value
                    Cod_Canal = row_data[21].value
                    Genero = row_data[22].value
                    PROM_Peso_Pie = row_data[23].value
                    Presentacion = row_data[24].value
                    C_Caliente = row_data[25].value
                    Grasa_Dorsal = row_data[26].value
                    RTO_PCC = row_data[27].value
                    C_Fria = row_data[28].value
                    RTO_PCF = row_data[29].value
                    Clasificacion = row_data[30].value
                    Merma = row_data[31].value
                    Magro = row_data[32].value
                    Cava = row_data[33].value
                    Tiempo_Cava = row_data[34].value
                    Clas_SEUROP = row_data[35].value
                    Clas_ABC = row_data[36].value
                    F_Remision = row_data[37].value
                    Destino_Cliente = row_data[38].value
                    Destino_Remision = row_data[39].value
                    Direccion_Remision = row_data[40].value
                    Mun_destino = row_data[41].value
                    Dep_destino = row_data[42].value
                    Remision = row_data[43].value
                    Tipo_Remision = row_data[44].value
                    Desposte = row_data[45].value
                    G_Invima = row_data[46].value
                    Placa_Furgon = row_data[47].value
                    Factura_Beneficio = row_data[48].value
                    Decomiso = row_data[49].value
                    Patologia = row_data[50].value
                    Observacion = row_data[51].value

                    Merma = Merma.replace('%', '') if Merma and '%' in Merma else Merma
                    RTO_PCC = RTO_PCC.replace('%', '') if RTO_PCC and '%' in RTO_PCC else RTO_PCC
                    RTO_PCF = RTO_PCF.replace('%', '') if RTO_PCF and '%' in RTO_PCF else RTO_PCF
                    Magro = Magro.replace('%', '') if Magro and '%' in Magro else Magro
                    
                    # Ejecutar una consulta SQL para insertar los datos en la tabla TRAZABILIDAD_OINC
                    cursor.execute(
                        'INSERT INTO trazabilidad_oinc (Mes, Año, Semana, Lote, Lote_Turn_Bene, Lote_Cod_Canal, Fase, F_Ingreso, F_Beneficio, F_Vencimiento, Solicitante, Propietario, Granja, Rem_Granja, Rem_Solicitante, Mun_Granja, Guia_ICA, Verificacion_ICA, Lab_IC, Registro_IC, Turno_Beneficio, Cod_Canal, Genero, PROM_Peso_Pie, Presentacion, C_Caliente, Grasa_Dorsal, RTO_PCC, C_Fria, RTO_PCF, Clasificacion, Merma, Magro, Cava, Tiempo_Cava, Clas_SEUROP, Clas_ABC, F_Remision, Destino_Cliente, Destino_Remision, Direccion_Remision, Mun_destino, Dep_destino, Remision, Tipo_Remision, Desposte, G_Invima, Placa_Furgon, Factura_Beneficio, Decomiso, Patologia, Observacion) VALUES (%s, %s,%s, %s, %s, %s,%s, %s, %s, %s,%s, %s,%s, %s, %s, %s,%s, %s, %s, %s,%s, %s,%s, %s, %s, %s,%s, %s, %s, %s,%s, %s,%s, %s, %s, %s,%s, %s, %s, %s,%s, %s,%s, %s, %s, %s,%s, %s, %s, %s, %s, %s)',
                        (Mes, Año, Semana, Lote, Lote_Turn_Bene, Lote_Cod_Canal, Fase, F_Ingreso, F_Beneficio, F_Vencimiento, Solicitante, Propietario, Granja, Rem_Granja, Rem_Solicitante, Mun_Granja, Guia_ICA, Verificacion_ICA, Lab_IC, Registro_IC, Turno_Beneficio, Cod_Canal, Genero, PROM_Peso_Pie, Presentacion, C_Caliente, Grasa_Dorsal, RTO_PCC, C_Fria, RTO_PCF, Clasificacion, Merma, Magro, Cava, Tiempo_Cava, Clas_SEUROP, Clas_ABC, F_Remision, Destino_Cliente, Destino_Remision, Direccion_Remision, Mun_destino, Dep_destino, Remision, Tipo_Remision, Desposte, G_Invima, Placa_Furgon, Factura_Beneficio, Decomiso, Patologia, Observacion)
                    )
                messages.success(request, 'Carga de datos en TRAZABILIDAD_OINC exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')


#-------- vista para el cargue de excel en PQRSF --------------------------------------------------------
@never_cache
@login_required
def cargar_excel_pqrsf(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_c
            with connections['b_c'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    PROCESO,TIPO,ESTADO_MOTIVO,CANTIDAD,CATEGORIA,TIEMPO_RESPUESTA,FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla PQRSF
                    cursor.execute(
                        'INSERT INTO pqrsf (PROCESO,TIPO,ESTADO_MOTIVO,CANTIDAD,CATEGORIA,TIEMPO_RESPUESTA,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s)',
                        (PROCESO.value,TIPO.value,ESTADO_MOTIVO.value,CANTIDAD.value,CATEGORIA.value,TIEMPO_RESPUESTA.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en PQRSF exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
# ----------------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#--------------------------------- CARGA DE   T.I ------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#-------- vista para el cargue de excel en Transformacion Digital --------------------------------------------------------
@login_required
def cargar_excel_transfordig(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ti
            with connections['b_ti'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    PROYECTO_ESTRATEGICO,CAPA_ARQUITECTURA,NOMBRE_PROYECTO,PESO_CAPA,PESO_PROYECTO_ESTRATEGICO,PORCENTAJE_AVANCE,PORCENTAJE_META,PORCENTAJE_META_PROYECTO,TAREAS_PROYECTO,TAREAS_PLANEADAS,TAREAS_EJECUTADAS,COSTO_PLANEADO,COSTO_EJECUTADO,FECHA_CORTE = row

                    # Ejecuta una consulta SQL para insertar los datos en la tabla TRANSFORMACION_DIGITAL
                    cursor.execute(
                        'INSERT INTO transformacion_digital (PROYECTO_ESTRATEGICO,CAPA_ARQUITECTURA,NOMBRE_PROYECTO,PESO_CAPA,PESO_PROYECTO_ESTRATEGICO,PORCENTAJE_AVANCE,PORCENTAJE_META,PORCENTAJE_META_PROYECTO,TAREAS_PROYECTO,TAREAS_PLANEADAS,TAREAS_EJECUTADAS,COSTO_PLANEADO,COSTO_EJECUTADO,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s)',
                        (PROYECTO_ESTRATEGICO.value,CAPA_ARQUITECTURA.value,NOMBRE_PROYECTO.value,PESO_CAPA.value,PESO_PROYECTO_ESTRATEGICO.value,PORCENTAJE_AVANCE.value,PORCENTAJE_META.value,PORCENTAJE_META_PROYECTO.value,TAREAS_PROYECTO.value,TAREAS_PLANEADAS.value,TAREAS_EJECUTADAS.value,COSTO_PLANEADO.value,COSTO_EJECUTADO.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en TRANSFORMACION_DIGITAL exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')

#-------- vista para el cargue de excel en Indicadores Economicos --------------------------------------------------------

@login_required
def cargar_excel_inideco(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = load_workbook(archivo_excel)
            ws = wb.active

            # Abre una conexión a la base de datos b_ti
            with connections['b_ti'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    valores = []
                    for cell in row:
                        if isinstance(cell.value, str):
                            valores.append(cell.value.upper())
                        elif isinstance(cell.value, int) or isinstance(cell.value, float):
                            valores.append(str(cell.value))
                        else:
                            valores.append(None)
                    
                    # Ejecuta una consulta SQL para insertar los datos en la tabla INDICADORES_ECONOMICOS
                    cursor.execute(
                        'INSERT INTO indicadores_economicos (INDICADOR,VALOR,FUENTE,LINK,FECHA_CORTE) VALUES (%s, %s, %s, %s, %s)',
                        tuple(valores)
                    )
                messages.success(request, 'Carga de datos en Indicadores Economicos exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#-------- vista para el cargue de excel en Transformacion Digital --------------------------------------------------------
@login_required
def cargar_excel_avantransfordig(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_ti
            with connections['b_ti'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    ACTIVIDAD,AVANCE,FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla TRANSFORMACION_DIGITAL
                    cursor.execute(
                        'INSERT INTO avance_transformacion_difital (ACTIVIDAD,AVANCE,FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s)',
                        (ACTIVIDAD.value,AVANCE.value,FECHA_CORTE.value,guid,usuario.username)
                    )
                messages.success(request, 'Carga de datos en avance transformacion digital exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
# ----------------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#--------------------------------- CARGA DE----Gestion Admin Financiera ------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
#-------- vista para el cargue de excel en  Compras Materia Prima ------------------------------------------------

@login_required
def cargar_excel_compramatprima(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_c
            with connections['b_gaf'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    MATERIA_PRIMA, COSTO_PROMEDIO, CANTIDAD_COMPRADA, DIAS_INVENTARIO, FECHA_CORTE = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla compras_materia_prima
                    cursor.execute(
                        'INSERT INTO compras_materia_prima (MATERIA_PRIMA, COSTO_PROMEDIO, CANTIDAD_COMPRADA, DIAS_INVENTARIO, FECHA_CORTE,GUID,USUARIO) VALUES (%s, %s, %s, %s, %s, %s, %s)',
                        (MATERIA_PRIMA.value , COSTO_PROMEDIO.value, CANTIDAD_COMPRADA.value, DIAS_INVENTARIO.value,FECHA_CORTE.value,guid, usuario.username)
                    )
                messages.success(request, 'Carga de datos en compras_materia_prima exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')


#-------- vista para el cargue de excel en oinc --------------------------------------------------------

# @login_required
# def cargar_excel_compramed(request):
#     if request.method == 'POST':
#         try:
#             usuario = request.user
            
#             archivo_excel = request.FILES['archivo_excel']
#             wb = load_workbook(archivo_excel)
#             ws = wb.active
            
#             guid = str(uuid4())

#             # Abre una conexión a la base de datos b_gaf
#             with connections['b_gaf'].cursor() as cursor:
#                 for row in ws.iter_rows(min_row=2):
#                     valores = []
#                     for cell in row:
#                         if isinstance(cell.value, str):
#                             valores.append(cell.value.upper())
#                         elif isinstance(cell.value, int) or isinstance(cell.value, float):
#                             valores.append(str(cell.value))
#                         else:
#                             valores.append(None)
#                     print(valores)
#                     # Ejecuta una consulta SQL para insertar los datos en la tabla compras_medicamentos
#                     cursor.execute(
#                         'INSERT INTO compras_medicamentos (VALOR, MEDICAMENTO, CLASIFICACION, CANTIDAD, TIPO, FECHA_CORTE) VALUES (%s, %s, %s, %s, %s, %s)',
#                        tuple(valores)
#                     )
#                 messages.success(request, 'Carga de datos en compras_medicamentos exitosa')
#         except KeyError:
#             messages.error(request, 'No se ha proporcionado un archivo Excel.')
#         except IntegrityError as e:
#             messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
#         except Exception as e:
#             messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
#         return redirect('home')
#     return render(request, '/home/')

@login_required
def cargar_excel_compramed(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user

            # Abre una conexión a la base de datos b_c
            with connections['b_gaf'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    print(row)
                    VALOR, MEDICAMENTO, CLASIFICACION, CANTIDAD, TIPO, FECHA_CORTE= row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla compras_medicamentos
                    cursor.execute(
                        'INSERT INTO compras_medicamentos (VALOR, MEDICAMENTO, CLASIFICACION, CANTIDAD, TIPO, FECHA_CORTE, GUID, USUARIO) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)',
                        (VALOR.value, MEDICAMENTO.value, CLASIFICACION.value, CANTIDAD.value, TIPO.value, FECHA_CORTE.value, guid, usuario.username)
                    )
                messages.success(request, 'Carga de datos en compras_medicamentos exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')


from django.db import connections

@login_required
def cargar_excel_preciocanal(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user

            # Abre una conexión a la base de datos dhc
            with connections['dhc'].cursor() as dhc_cursor:
                # Obtener todos los NITs existentes en la tabla clientes de dhc
                dhc_cursor.execute('''SELECT nit FROM dhc.clientes''')
                nits_existentes = [row[0] for row in dhc_cursor.fetchall()]

            # Conjunto para almacenar NITs vistos en el archivo
            nits_vistos = set()

            # Abre una conexión a la base de datos b_gaf
            with connections['b_gaf'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    NIT, CLIENTE, ZONA, VALOR, SEMANA = row
                    
                    # Verificar si el NIT existe en la lista de NITs existentes
                    if NIT.value is not None and int(NIT.value) not in map(int, nits_existentes):
                        messages.error(request, f'Error: El NIT {NIT.value} no existe en la base de datos de clientes.')
                        return redirect('home')

                    # # Verificar si el NIT está repetido en el archivo subido
                    # if NIT.value in nits_vistos:
                    #     messages.error(request, f'Error: El NIT {NIT.value} está repetido en el archivo.')
                    #     return redirect('home')
                    
                    # Añadir el NIT al conjunto de NITs vistos
                    nits_vistos.add(NIT.value)

                    # Ejecuta una consulta SQL para insertar los datos en la tabla precio_canales_semana
                    cursor.execute(
                        'INSERT INTO precio_canales_semana (NIT, CLIENTE, ZONA, VALOR,SEMANA, GUID, USUARIO) VALUES (%s, %s, %s, %s, %s, %s, %s)',
                        (NIT.value, CLIENTE.value, ZONA.value, VALOR.value, SEMANA.value, guid, usuario.username)
                    )
                messages.success(request, 'Carga de datos en precio canales exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')






@login_required
def cargar_excel_clientes(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user

            # Abre una conexión a la base de datos b_c
            with connections['dhc'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    
                    NIT,RAZON_SOCIAL,CUPO,DIRECCION_SEDE_PRINCIPAL,DIRECCION_EXPENDIO,ID_CLASIFICACION,ID_MUNICIPIO,ID_DEPARTAMENTO,ID_REGION,ID_VENDEDOR,ID_SEGMENTO,ID_linea_negocio, = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla precio_canales_semana
                    cursor.execute(
                        'INSERT INTO clientes (NIT, RAZON_SOCIAL, CUPO, DIRECCION_SEDE_PRINCIPAL, DIRECCION_EXPENDIO, ID_CLASIFICACION, ID_MUNICIPIO, ID_DEPARTAMENTO, ID_REGION, ID_VENDEDOR, ID_SEGMENTO, ID_linea_negocio, GUID, USUARIO) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (NIT.value, RAZON_SOCIAL.value, CUPO.value, DIRECCION_SEDE_PRINCIPAL.value, DIRECCION_EXPENDIO.value, ID_CLASIFICACION.value, ID_MUNICIPIO.value, ID_DEPARTAMENTO.value, ID_REGION.value, ID_VENDEDOR.value, ID_SEGMENTO.value, ID_linea_negocio.value, guid, usuario.username)
                    )

                    print(row)
                messages.success(request, 'Carga de datos en clientes exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en Evolucion Precio Canal----------------------------------------------
@never_cache
@login_required
def cargar_excel_evolucion_precio_canal(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gc
            with connections['b_gaf'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    FECHA_REPORTE,PRECIO_PROM_CANAL_FRIA,PRECIO_PROM_KG_GRANJA,CANTIDAD_CERDOS_VENDIDOS = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla compromiso_mes
                    cursor.execute(
                        'INSERT INTO evolucion_precio_canal ( FECHA_REPORTE,PRECIO_PROM_CANAL_FRIA,PRECIO_PROM_KG_GRANJA,CANTIDAD_CERDOS_VENDIDOS,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s)',
                        ( FECHA_REPORTE.value,PRECIO_PROM_CANAL_FRIA.value,PRECIO_PROM_KG_GRANJA.value,CANTIDAD_CERDOS_VENDIDOS.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en evolucion_precio_canal exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en costo_kg_producido_kg_vendido----------------------------------------------
@never_cache
@login_required
def cargar_excel_costo_kg_producido_kg_vendido(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gc
            with connections['b_gaf'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):

                    FECHA_REPORTE,TONELADAS_PRODUCIDAS,COSTO_PROMEDIO_PRO,TONELADAS_VENDIDAS,VALOR_PROMEDIO_VENTA_KG,MARGEN_BRUTO = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla compromiso_mes
                    cursor.execute(
                        'INSERT INTO costo_kg_producido_kg_vendido ( FECHA_REPORTE,TONELADAS_PRODUCIDAS,COSTO_PROMEDIO_PRO,TONELADAS_VENDIDAS,VALOR_PROMEDIO_VENTA_KG,MARGEN_BRUTO,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s, %s)',
                        ( FECHA_REPORTE.value,TONELADAS_PRODUCIDAS.value,COSTO_PROMEDIO_PRO.value,TONELADAS_VENDIDAS.value,VALOR_PROMEDIO_VENTA_KG.value,MARGEN_BRUTO.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en costo_kg_producido_kg_vendido exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')
#------ vista para el cargue de excel en indicadores_economicos----------------------------------------------
@never_cache
@login_required
def cargar_excel_indicadores_economicos(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gc
            with connections['b_gaf'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    FECHA_REPORTE,INDICADOR,VALOR,FUENTE,LINK = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla compromiso_mes
                    cursor.execute(
                        'INSERT INTO indicadores_economicos (FECHA_REPORTE,INDICADOR,VALOR,FUENTE,LINK,GUID,USUARIO) VALUES (%s, %s,%s, %s, %s, %s, %s)',
                        ( FECHA_REPORTE.value,INDICADOR.value,VALOR.value,FUENTE.value,LINK.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en indicadores_economicos exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')

@never_cache
@login_required
def cargar_excel_costopromediomp(request):
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            guid = str(uuid4())
            usuario = request.user
            # Abre una conexión a la base de datos b_gc
            with connections['b_gaf'].cursor() as cursor:
                for row in ws.iter_rows(min_row=2):
                    fecha_reporte,materia_prima,valor_promedio = row
                    # Ejecuta una consulta SQL para insertar los datos en la tabla compromiso_mes
                    cursor.execute(
                        'INSERT INTO costo_promedio_mp (fecha_reporte,materia_prima,valor_promedio,guid,usuario) VALUES (%s, %s,%s, %s, %s)',
                        ( fecha_reporte.value,materia_prima.value,valor_promedio.value,guid,usuario.username)
                    )
            messages.success(request, 'Carga de datos en indicadores_economicos exitosa')
        except KeyError:
            messages.error(request, 'No se ha proporcionado un archivo Excel.')
        except IntegrityError as e:
            messages.error(request, f'Error al insertar datos en la base de datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Se ha producido un error inesperado: {str(e)}')
        return redirect('home')
    return render(request, '/home/')

# ------------------------------------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------------------------------------
#----------------------------Funciones   internas utilizadas por los html-------------------------------------------------------
# ------------------------------------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------------------------------------

@never_cache
@login_required
def reproved(request):
    with connections['b_ca'].cursor() as cursor:
        cursor.execute('SELECT granja, mes, semana, cantidad_cerdos FROM compromiso_mes')
        compromisos = cursor.fetchall()

    data = [{'granja': granja, 'mes': mes, 'semana': semana, 'cantidad_cerdos': cantidad_cerdos} for granja, mes, semana, cantidad_cerdos in compromisos]

    response = JsonResponse({'data': data})
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response

logger = logging.getLogger(__name__)



def repfinan(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
   
    
    with connections['b_gaf'].cursor() as cursor:
        cursor.execute('''
            SELECT Fecha_transformacion,Unidades,Peso_canal_fria,Consecutivo_Cercafe,Codigo_granja,Remision,Valor,Cliente,Planta_Beneficio,Granja,Nit_asociado,Asociado,Grupo_Granja,Retencion,Valor_a_pagar_asociado,Valor_kilo,id
            FROM b_gaf.operacion_desposte
            WHERE Fecha_transformacion BETWEEN %s AND %s
        ''', [start_date, end_date])
        compromisos = cursor.fetchall()

    # Loguear los datos recuperados
    logger.info(compromisos)

    data = [{'Fecha_transformacion': Fecha_transformacion, 'Unidades': Unidades, 'Peso_canal_fria': Peso_canal_fria, 'Consecutivo_Cercafe': Consecutivo_Cercafe, 'Codigo_granja': Codigo_granja, 'Remision': Remision, 'Valor': Valor, 'Cliente': Cliente, 'Planta_Beneficio': Planta_Beneficio, 'Granja': Granja, 'Nit_asociado': Nit_asociado, 'Asociado': Asociado, 'Grupo_Granja': Grupo_Granja, 'Retencion': Retencion, 'Valor_a_pagar_asociado': Valor_a_pagar_asociado, 'Valor_kilo': Valor_kilo, 'id': id } for Fecha_transformacion, Unidades, Peso_canal_fria, Consecutivo_Cercafe, Codigo_granja, Remision, Valor, Cliente, Planta_Beneficio, Granja, Nit_asociado, Asociado, Grupo_Granja, Retencion, Valor_a_pagar_asociado, Valor_kilo, id in compromisos]

    return JsonResponse({'data': data})


def get_filtered_data(start_date, end_date):
    with connections['b_gaf'].cursor() as cursor:
        cursor.execute('''
            SELECT fecha_transformacion,Unidades,Peso_canal_fria,Consecutivo_Cercafe,Codigo_granja,Remision,Valor,Cliente,Planta_Beneficio,Granja,Nit_asociado,Asociado,Grupo_Granja,Retencion,Valor_a_pagar_asociado,Valor_kilo,id
            FROM b_gaf.operacion_desposte
            WHERE fecha_transformacion BETWEEN %s AND %s
        ''', [start_date, end_date])
        compromisos = cursor.fetchall()

    # Loguear los datos recuperados
    logger.info(compromisos)

    return compromisos


def export_pdf(request):
    # Obtener los datos para exportar
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Obtener los datos filtrados
    compromisos = get_filtered_data(start_date, end_date)

    # Crear el HTML con los datos filtrados
    html = '<html><body><table><thead><tr><th>Fecha Transformación</th><th>Unidades</th><th>Peso Canal Fría</th><th>Consecutivo Cercafe</th><th>Código Granja</th><th>Remisión</th><th>Valor</th><th>Cliente</th><th>Planta Beneficio</th><th>Granja</th><th>Nit Asociado</th><th>Asociado</th><th>Grupo Granja</th><th>Retención</th><th>Valor a Pagar Asociado</th><th>Valor Kilo</th></tr></thead><tbody>'
    for compromiso in compromisos:
        html += '<tr>'
        for value in compromiso:
            html += '<td>' + str(value) + '</td>'
        html += '</tr>'
    html += '</tbody></table></body></html>'

    # Convertir el HTML a PDF
    pdf = pdfkit.from_string(html, False, configuration=pdfkit.configuration(wkhtmltopdf='C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe'))

    # Enviar el PDF como respuesta de descarga
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'
    return response




def export_excel(request):
    # Obtener los datos para exportar
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Obtener los datos filtrados
    compromisos = get_filtered_data(start_date, end_date)

    # Obtener el directorio de descargas del usuario actual
    downloads_dir = os.path.join(settings.BASE_DIR, 'tmp')

    # Crear el archivo Excel en el directorio de descargas
    filename = 'reporte.xlsx'
    file_path = os.path.join(downloads_dir, filename)

    # Escribir los datos en el archivo Excel
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()

    # Escribir los encabezados
    headers = ['Fecha Transformación', 'Unidades', 'Peso Canal Fría', 'Consecutivo_Cercafe', 'Código Granja', 'Remisión', 'Valor', 'Cliente', 'Planta Beneficio', 'Granja', 'Nit Asociado', 'Asociado', 'Grupo Granja', 'Retención', 'Valor a Pagar Asociado', 'Valor Kilo','id']
    for i, header in enumerate(headers):
        worksheet.write(0, i, header)

    # Obtener el índice de la columna 'id'
    id_column_index = headers.index('id') if 'id' in headers else None

    # Escribir los datos
    for row, compromiso in enumerate(compromisos, start=1):
        for col, value in enumerate(compromiso):
            # Formatear la fecha como un string
            if isinstance(value, datetime.date):
                value = value.strftime('%Y-%m-%d')
            # Verificar si la columna actual no es 'id'
            if id_column_index is None or col != id_column_index:
                worksheet.write(row, col, value)

    # Cerrar el archivo Excel
    workbook.close()

    # Enviar el archivo Excel como respuesta de descarga
    with open(file_path, 'rb') as file:
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


    
    
@csrf_protect
def save_changes(request):
    if request.method == 'POST':
        # Obtener los datos del formulario
        id = request.POST.get('id')
        newValue = request.POST.get('newValue')

        # Validar los datos
        if not id.isdigit():
            return JsonResponse({'success': False, 'error': 'El ID debe ser un número entero válido'})

        # Realizar la actualización en la base de datos
        try:
            # Actualizar el campo 'Valor Kilo'
            with connections['b_gaf'].cursor() as cursor:
                cursor.execute('''
                    UPDATE b_gaf.operacion_desposte
                    SET valor_kilo = %s
                    WHERE id = %s
                ''', [newValue, id])
            # Devolver una respuesta de éxito
            return JsonResponse({'success': True})
        except Exception as e:
            # Devolver una respuesta de error si ocurre algún error
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        # Devolver una respuesta de error si la solicitud no es POST
        return JsonResponse({'success': False, 'error': 'Método de solicitud no permitido'})
    
def get_filtered_data_by_group(start_date, end_date, selected_group):
    with connections['b_gaf'].cursor() as cursor:
        cursor.execute('''
            SELECT Asociado,Granja,Cliente,Unidades,Peso_canal_fria,Valor_kilo,Valor,Retencion,Valor_a_pagar_asociado
            FROM b_gaf.operacion_desposte
            WHERE fecha_transformacion BETWEEN %s AND %s AND grupo_granja = %s
        ''', [start_date, end_date, selected_group])
        compromisos = cursor.fetchall()

    # Loguear los datos recuperados
    logger.info(compromisos)

    return compromisos

from collections import defaultdict
from django.http import HttpResponse 

def generate_excel_report(request):
    import xlsxwriter

    # Obtener los datos para exportar
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_group = request.GET.get('selected_group')

    # Obtener los datos filtrados para el grupo seleccionado
    compromisos = get_filtered_data_by_group(start_date, end_date, selected_group)

    # Convertir las tuplas en listas para que puedan ser ordenadas
    compromisos = list(compromisos)

    # Ordenar los compromisos por asociado, granja y cliente
    compromisos.sort(key=lambda x: (x[0], x[1], x[2]))  

    # Crear el archivo Excel
    workbook = xlsxwriter.Workbook('reporte_grupo_' + selected_group + '.xlsx')
    worksheet = workbook.add_worksheet()

    # Formato para las celdas que cumplen la condición
    formato_azul = workbook.add_format({'bg_color': '#9BC2E6', 'bold': True})  # Encabezados en azul y en negrita
    bordeado = workbook.add_format({'border': 1})  # Formato para bordes
    formato_amarillo = workbook.add_format({'bg_color': 'yellow'})

    # Escribir los encabezados
    headers = ['Asociado', 'Granja', 'Cliente', 'Unidades', 'Peso Canal', 'Valor Kilo', 'Valor a facturar', 'Retención',
               'Valor a Pagar Asociado']
    for i, header in enumerate(headers):
        worksheet.write(0, i, header, bordeado)
        worksheet.write(0, i, header, formato_azul) 
        
    # Ajustar automáticamente el ancho de las columnas al texto de los encabezados
    for i, header in enumerate(headers):
        worksheet.set_column(i, i, len(header) + 2)  # Ajustar el ancho de la columna al tamaño del encabezado más un margen
        worksheet.freeze_panes(1, 0)
    # Escribir los datos
    current_row = 1
    current_granja = None
    granja_total_values = [0, 0, 0, 0, 0, 0, 0, 0, 0]  # Inicializar los valores totales para cada columna
    consolidado_unidades = 0
    consolidado_kilos = 0
    consolidado_valor_total = 0
    for compromiso in compromisos:
        if current_granja is None or current_granja != compromiso[1]:
            # Si la granja cambió, escribir los totales y reiniciar los valores totales
            if current_row > 1:
                current_row += 1
                worksheet.write(current_row, 0, "")
                worksheet.write(current_row, 1, "TOTAL :")
                # Escribir los totales de la granja anterior con bordes y formato amarillo
                for col, total_value in enumerate(granja_total_values[3:], start=3):
                    worksheet.write(current_row, col, total_value, bordeado)
                    worksheet.write(current_row, col, total_value, formato_amarillo)  # Aplicar formato amarillo a las celdas de estas columnas
                      # Aplicar bordes
                current_row += 2  # Saltar dos filas para separar los bloques
            current_granja = compromiso[1]
            granja_total_values = [0, 0, 0, 0, 0, 0, 0, 0, 0]  # Reiniciar los valores totales

        # Escribir los valores en las columnas correspondientes
        for col, value in enumerate(compromiso):
            worksheet.write(current_row, col, value, bordeado)  # Aplicar bordes a todas las celdas
            if col in [3, 4, 6, 7, 8]:  # Columnas para las cuales calcular los totales
                granja_total_values[col] += float(value.replace(',', '.')) if isinstance(value, str) else value
        
        # Actualizar los totales consolidados
        consolidado_unidades += compromiso[3]
        consolidado_kilos += float(compromiso[4].replace(',', '.')) if isinstance(compromiso[4], str) else compromiso[4]
        consolidado_valor_total += float(compromiso[6].replace(',', '.')) if isinstance(compromiso[6], str) else compromiso[6]
        current_row += 1

    # Escribir los totales para la última granja con bordes y formato amarillo
    if granja_total_values:
        current_row += 1
        worksheet.write(current_row, 0, "")
        worksheet.write(current_row, 5, "")
        worksheet.write(current_row, 1, "TOTAL :")
        # Escribir los totales de la última granja con bordes y formato amarillo
        for col, total_value in enumerate(granja_total_values[3:], start=3):
            worksheet.write(current_row, col, total_value, bordeado)
            worksheet.write(current_row, col, total_value, formato_amarillo)  # Aplicar formato amarillo a las celdas de estas columnas
              # Aplicar bordes

    for i, header in enumerate(headers):
        worksheet.set_column(i, i, len(header) + 2) 

    consolidado_data = {}
    total_unidades = 0
    total_kilos = 0
    total_valor_total = 0

    for compromiso in compromisos:
        # Tu código para escribir los datos del bloque

        # Acumular información para el cuadro consolidado
        asociado = compromiso[0]
        granja = compromiso[1]
        unidades = compromiso[3]
        kilos = float(compromiso[4].replace(',', '.')) if isinstance(compromiso[4], str) else compromiso[4]
        valor_total = float(compromiso[6].replace(',', '.')) if isinstance(compromiso[6], str) else compromiso[6]

        if asociado not in consolidado_data:
            consolidado_data[asociado] = {}
        if granja not in consolidado_data[asociado]:
            consolidado_data[asociado][granja] = {'unidades': 0, 'kilos': 0, 'valor_total': 0}

        consolidado_data[asociado][granja]['unidades'] += unidades
        consolidado_data[asociado][granja]['kilos'] += kilos
        consolidado_data[asociado][granja]['valor_total'] += valor_total

        # Acumular totales globales
        total_unidades += unidades
        total_kilos += kilos
        total_valor_total += valor_total

    
    cuadro_data = [["","CONSOLIDADO","","",""], ["Asociado", "Granja", "Unidades", "Kilos", "Valor Total"]]
    row_start = current_row + 3
    for asociado, granjas in consolidado_data.items():
        for granja, datos in granjas.items():
            cuadro_data.append([asociado, granja, datos['unidades'], datos['kilos'], "${:,.2f}".format(datos['valor_total'])])

    # Agregar la fila de totales al final del cuadro consolidado
    worksheet.write(current_row, 0, "")
    cuadro_data.append(['', 'TOTAL :', total_unidades, total_kilos, "${:,.2f}".format(total_valor_total)])

    for row_num, row_data in enumerate(cuadro_data):
        for col_num, col_data in enumerate(row_data):
            worksheet.write(row_start + row_num, col_num, col_data, bordeado)
            if row_num == 0 or row_num == 1:  # Formato para las primeras dos filas
                worksheet.write(row_start + row_num, col_num, col_data, formato_azul)
            elif row_num == len(cuadro_data) - 1:  # Formato para la última fila
                worksheet.write(row_start + row_num, col_num, col_data, formato_amarillo)

                    


    # Cerrar el archivo Excel
    workbook.close()

    # Enviar el archivo Excel como respuesta
    with open('reporte_grupo_' + selected_group + '.xlsx', 'rb') as file:
        response = HttpResponse(file.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_grupo_' + selected_group + '.xlsx"'
        return response  


##-------------------- PARAMETROS GENERALES-----------------------

def grupos_asociados(request):
    with connections['dhc'].cursor() as cursor:
        cursor.execute('''SELECT grupo_asociado FROM dhc.grupo_asociado''')
        grupos_asociados = [row[0] for row in cursor.fetchall()]
        print(grupos_asociados)  
    return grupos_asociados

def granjas(request):
    with connections['dhc'].cursor() as cursor:
        cursor.execute('''SELECT id, GRANJAS FROM dhc.granjas''')  
        granjas = [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]  
    
    return JsonResponse({'granjas': granjas})

def sitio(request):
    with connections['dhc'].cursor() as cursor:
        cursor.execute('''SELECT id, nombre_sitio FROM dhc.p_sitio_ps''')  
        sitio = [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]  
    
    return JsonResponse({'sitio': sitio})


def caracteristicas(request):
    with connections['dhc'].cursor() as cursor:
        cursor.execute('''SELECT id, ncaracteristica FROM dhc.p_caracteristicas''')  
        caracteristicas = [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]  
    
    return JsonResponse({'caracteristicas': caracteristicas})

def genero(request):
    with connections['dhc'].cursor() as cursor:
        cursor.execute('''SELECT id, ngenero FROM dhc.p_genero''')  
        genero = [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]  
    
    return JsonResponse({'genero': genero})

def frigorificos(request):
    with connections['dhc'].cursor() as cursor:
        cursor.execute('''SELECT id, nombre FROM dhc.frigorificos''')  
        frigorificos = [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]  
    
    return JsonResponse({'frigorificos': frigorificos})

def placas(request):
    with connections['dhc'].cursor() as cursor:
        cursor.execute('''SELECT id, placa FROM dhc.placas''')  
        placas = [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]  
    
    return JsonResponse({'placas': placas})

def conductores(request):
    with connections['dhc'].cursor() as cursor:
        cursor.execute('''SELECT id, conductor FROM dhc.Conductores''')  
        conductor = [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]  
    
    return JsonResponse({'conductor': conductor})


#---------------- TABLAS DE REPORTES G COMERCIAL------------------------------------------
#---Define La Vista Rep-gestion comercial----
@never_cache
@login_required
def repgcomercial(request):
    clientes = tablarepclient(request)
    ventas = tablarepventas(request) 
    return render(request, 'report_gcomercial.html', {'clientes_act': clientes, 'repventas': ventas})

def tablarepclient(request):
    with connections['b_gc'].cursor() as cursor:
        cursor.execute('''SELECT FECHA_CORTE,CANTIDAD_CLIENTES,ZONA_CLIENTE,KG_FACTURADOS,DINERO_APORTADO,ESTADO_CLIENTE FROM b_gc.clientes_activos
                          WHERE GUID = (SELECT MAX(GUID) FROM b_gc.clientes_activos) ''')
        clientes_act = cursor.fetchall()   
    return clientes_act
def tablarepventas(request):
    with connections['b_gc'].cursor() as cursor:
        cursor.execute('''SELECT FECHA_CORTE,LINEA_NEGOCIO,PRESUPUESTO_UNIDADES,PRESUPUESTO_KG,UNIDADES_VENDIDAS,KG_VENDIDO,VALOR_VENTA,PRESUPUESTO_VENTA FROM b_gc.ventas 
                          WHERE GUID = (SELECT MAX(GUID) FROM b_gc.ventas)''')
        repventas = cursor.fetchall()   
    return repventas
#---------------- TABLAS DE REPORTES G TECNICA------------------------------------------
@never_cache
@login_required
def repgtecnica(request):
    abhembras = tablarepabhembras(request)
    fortuitos = tablarepfortuitos(request) 
    kgvendidos = tablarepkgvendidos(request) 
    pfinalcon = tablareppfinalcon(request) 
    prohembras = tablarepprohembras(request) 
    tecnicacia = tablareptecnicacia(request) 
    return render(request, 'report_gtecnica.html', {'abhembras': abhembras,'fortuitos':fortuitos,'kgvendidos':kgvendidos,'pfinalcon':pfinalcon,'prohembras':prohembras,'tecnicacia':tecnicacia})

def tablarepabhembras(request):
    with connections['b_gt'].cursor() as cursor:
        cursor.execute('''SELECT GRANJA,CANTIDAD_ENTREGADA,PORCENTAJE_CUMPLIMIENTO,FECHA_CORTE FROM b_gt.abastecimiento_hembras
                        WHERE GUID = (SELECT MIN(GUID) FROM b_gt.abastecimiento_hembras)''')
        abhembras = cursor.fetchall()   
    return abhembras
def tablarepfortuitos(request):
    with connections['b_gt'].cursor() as cursor:
        cursor.execute('''SELECT FECHA_CORTE,PLANTA,GRANJA,CANTIDAD_MUERTE_TRANSPORTE,CANTIDAD_MUERTE_REPOSO,AGITADOS,LESIONADOS,RETOMAS,TOTAL FROM b_gt.fortuitos3
                       WHERE GUID = (SELECT MAX(GUID) FROM b_gt.fortuitos3)''')
        fortuitos = cursor.fetchall()   
    return fortuitos
def tablarepkgvendidos(request):
    with connections['b_gt'].cursor() as cursor:
        cursor.execute('''SELECT GRANJA,KG_V_H_A,ASOCIADO,FECHA_CORTE FROM b_gt.kg_vendidos_hembra WHERE GUID = (SELECT MAX(GUID) FROM b_gt.kg_vendidos_hembra)''')
        kgvendidos = cursor.fetchall()   
    return kgvendidos
def tablareppfinalcon(request):
    with connections['b_gt'].cursor() as cursor:
        cursor.execute('''SELECT GRANJA,PESO,META_PESO,CONVERSION_META,CONVERSION,FECHA_CORTE FROM b_gt.peso_final_conversion WHERE GUID = (SELECT MAX(GUID) FROM b_gt.peso_final_conversion)''')
        pfinalcon = cursor.fetchall()   
    return pfinalcon
def tablarepprohembras(request):
    with connections['b_gt'].cursor() as cursor:
        cursor.execute('''SELECT PARTOS,TASA_PARTOS,CUMPLIMIENTO_PROYECTADO,CUMPLIMIENTO_REAL,AÑO_SERVICIO,OBSERVACIONES,FECHA_CORTE,CANTIDAD_PROYECCION_HEMBRAS,CANTIDAD_REAL FROM b_gt.proyeccion_hembras WHERE GUID = (SELECT MAX(GUID) FROM b_gt.proyeccion_hembras)''')
        prohembras = cursor.fetchall()   
    return prohembras
def tablareptecnicacia(request):
    with connections['b_gt'].cursor() as cursor:
        cursor.execute('''SELECT LINEA_GENETICA,CANTIDAD_MACHOS,PORCENTAJE_DISTRIBUCION_MACHOS,CANTIDAD_DESECHADO,PORCENTAJE_DESCECHADO,DOSIS_PRODUCIDAS,DOSIS_VENDIDAS,PROMEDIO_MORFOLOGIA,OBSERVACION,FECHA_CORTE FROM b_gt.tecnica_cia WHERE GUID = (SELECT MAX(GUID) FROM b_gt.tecnica_cia)''')
        tecnicacia = cursor.fetchall()   
    return tecnicacia
#---------------- TABLAS DE REPORTES CADENA DE ABASTECIMIENTO------------------------------------------
@never_cache
@login_required
def repcadabastecimiento(request):
    compgranja = tablarepcompgranja(request)
    disposemana = tablarepdisposem(request)
    cerdosbenef = tablarepcerdosbenef(request) 
    comparativopl = tablarepcomparativopl(request) 
    costodespo = tablarepcostodespo(request) 
    kgbenef = tablarepkgbenef(request) 
    kgdespos = tablarepkgdespos(request) 
    particortes = tablarepparticortes(request) 
    toneladasimport = tablareptoneladasimport(request) 
    return render(request, 'report_cadabastecimiento.html', {'compgranja': compgranja,'disposemana': disposemana,'cerdosbenef':cerdosbenef,'comparativopl':comparativopl,'costodespo':costodespo,'kgbenef':kgbenef,'kgdespos':kgdespos,'particortes':particortes,'toneladasimport':toneladasimport})

def tablarepcompgranja(request):
    with connections['b_ca'].cursor() as cursor:
        cursor.execute('''select  granja,mes,semana,cantidad_cerdos,año from b_ca.compromiso_mes''')
        compgranja = cursor.fetchall()
      
    return compgranja

def tablarepdisposem(request):
    with connections['b_ca'].cursor() as cursor:
        cursor.execute('''select  granja,mes,semana,cantidad_cerdos,año from b_ca.disponibilidad_semanal''')
        disposemana = cursor.fetchall()
    return disposemana
def tablarepcerdosbenef(request):
    with connections['b_ca'].cursor() as cursor:
        cursor.execute('''SELECT CER_BENEF_COLOMBIA,CER_BENEF_EJE_CAFETERO,PARTICIPACION_EJE_CAFETERO,CER_BENEF_CERCAFE,PARTICIPACION_EJE_CAF_CERCAFE,PARTICIPACION_NACIONAL_CERCAFE,FECHA_CORTE FROM b_ca.prod_carnica_cerdos_beneficiados ''')
        cerdosbenef = cursor.fetchall()   
    return cerdosbenef
def tablarepcomparativopl(request):
    with connections['b_ca'].cursor() as cursor:
        cursor.execute('''SELECT PARAMETRO,VALOR,EMPRESA,FECHA_CORTE FROM b_ca.prod_carnica_comparativo_plantas''')
        comparativopl = cursor.fetchall()   
    return comparativopl
def tablarepcostodespo(request):
    with connections['b_ca'].cursor() as cursor:
        cursor.execute('''SELECT TIPO_CLIENTE,NUM_CERDOS_DESPOSTADOS,KG_DESPOSTADOS,PESO_PROM_CERDOS,PRECIO_PROM_KG,COSTO_MATERIA_PRIMA,COSTO_MAQUILA,COSTO_KG_MAQUILADO,MAQUILA_SIN_MP,FECHA_CORTE FROM b_ca.prod_carnica_costo_desposte''')
        costodespo = cursor.fetchall()   
    return costodespo
def tablarepkgbenef(request):
    with connections['b_ca'].cursor() as cursor:
        cursor.execute('''SELECT CER_BENEF_COLOMBIA,CER_BENEF_EJE_CAFETERO,PARTICIPACION_EJE_CAFETERO,CER_BENEF_CERCAFE,PARTICIPACION_EJE_CAF_CERCAFE,PARTICIPACION_NACIONAL_CERCAFE,PESO_CF_NACIONAL,PESO_EJE_CAFETERO,PESO_CF_CERCAFE,KG_NACIONAL,KG_EJE_CAFETERO,KG_CERCAFE,FECHA_CORTE FROM b_ca.prod_carnica_kg_beneficio ''')
        kgbenef = cursor.fetchall()   
    return kgbenef
def tablarepkgdespos(request):
    with connections['b_ca'].cursor() as cursor:
        cursor.execute('''SELECT KG_PRODUCIDOS_CERCAFE,KG_DESPOSTADOS_CERCAFE,PORCENTAJE_PARTICIPACION,TRIMESTRE_2022_CERCAFE,TRIMESTRE_2022_DESPOSTE,TRIMESTRE_2023_CERCAFE,TRIMESTRE_2023_DESPOSTE,CERCIMIENTO_22_23,FECHA_CORTE FROM b_ca.prod_carnica_kg_despostados ''')
        kgdespos = cursor.fetchall()   
    return kgdespos
def tablarepparticortes(request):
    with connections['b_ca'].cursor() as cursor:
        cursor.execute('''SELECT CORTE,PORCENTAJE_PARTICIPACION,PORCENTAJE_META,PESO_PROM_CANAL,CANTIDAD_CANALES,FECHA_CORTE FROM b_ca.prod_carnica_participacion_cortes ''')
        particortes = cursor.fetchall()   
    return particortes
def tablareptoneladasimport(request):
    with connections['b_ca'].cursor() as cursor:
        cursor.execute('''SELECT CER_BENEF_COLOMBIA,TON_BENEF_COLOMBIA,TON_IMPORT_COLOMBIA,CERDOS_IMPORTADOS,ENE_FEB_22_TON_BENEF,ENE_FEB_23_TON_BENEF,CRECIMIENTO_22_23,ENE_FEB_MAR_22_TON_IMPORT,ENE_FEB_MAR_23_TON_IMPORT,CRECIMIENTO_OMPORT_22_23,FECHA_CORTE FROM b_ca.prod_carnica_ton_importadas ''')
        toneladasimport = cursor.fetchall()   
    return toneladasimport

#---------------- TABLAS DE REPORTES  ALIMENTO BALANCEADO ------------------------------------------
@never_cache
@login_required
def repplantaalibal(request):
    plantaalib = tablarepplantab(request)
    return render(request, 'report_galimento.html', {'plantaalib': plantaalib})

def tablarepplantab(request):
    with connections['b_gab'].cursor() as cursor:
        cursor.execute('''SELECT TONELADAS_PRODUCIDAS_MES,TONELADAS_PRESUPUESTO_MES,PORCENTAJE_VARIACION_MES,PORCENTAJE_CUMPLIMIENTO_MES,OBSERVACION_VARIACION,PORCENTAJE_BULTO_MES,PORCENTAJE_GRANEL_MES,SACK_OFF,PORCENTAJE_OTIF,OBSERVACION_OTIF,PRESUPUESTO_MO_CIF,MO_CIF,TIEMPO_MUERTO,COSTO_TIEMPO_MUERTO,OBSERVACION_TIEMPO_MUERTO,FECHA_CORTE FROM b_gab.planta_alimentos_balanceados ''')
        plantaalib = cursor.fetchall()   
    return plantaalib

#---------------- TABLAS DE REPORTES CALIDAD------------------------------------------
@never_cache
@login_required
def repcalidad(request):
    avancepro = tablarepavancepro(request)
    calidadpla = tablarepcalidadpla(request) 
    causadesvia = tablarepcausadesvia(request) 
    pqrsf = tablareppqrsf(request) 
    return render(request, 'report_calidad.html', {'avancepro': avancepro,'calidadpla':calidadpla,'causadesvia':causadesvia,'pqrsf':pqrsf})

def tablarepavancepro(request):
    with connections['b_c'].cursor() as cursor:
        cursor.execute('''SELECT TIPO,PROCESO,DETALLE_PROCESO,AVANCE,META,FECHA_CORTE FROM b_c.avance_proceso WHERE GUID = (SELECT MIN(GUID) FROM b_c.avance_proceso) ''')
        avancepro = cursor.fetchall()   
    return avancepro
def tablarepcalidadpla(request):
    with connections['b_c'].cursor() as cursor:
        cursor.execute('''SELECT PORCENTAJE_DESVIACIONES_CALIDAD,TONELADAS_REPROCESADAS,TONELADAS_LIBERADAS_CONCESION,PORCENTAJE_RETENCION,PORCENTAJE_MEZCLA,PORCENTAJE_DURABILIDAD,PORCENTAJE_FINOS,PORCENTAJE_FORMULACION,CUMPLIMIENTO_BPM,FECHA_CORTE FROM b_c.calidad_planta WHERE GUID = (SELECT MIN(GUID) FROM b_c.calidad_planta)''')
        calidadpla = cursor.fetchall()   
    return calidadpla
def tablarepcausadesvia(request):
    with connections['b_c'].cursor() as cursor:
        cursor.execute('''SELECT CAUSA,PLAN_ACCION,TON_REPROCESADAS,FECHA_CORTE FROM b_c.causas_desviaciones''')
        causadesvia = cursor.fetchall()   
    return causadesvia
def tablareppqrsf(request):
    with connections['b_c'].cursor() as cursor:
        cursor.execute('''SELECT PROCESO,TIPO,ESTADO_MOTIVO,CANTIDAD,CATEGORIA,TIEMPO_RESPUESTA,FECHA_CORTE FROM b_c.pqrsf''')
        pqrsf = cursor.fetchall()   
    return pqrsf

#---------------- TABLAS DE REPORTES GESTION ADMINISTRATIVA Y FINANCIERA------------------------------------------
@never_cache
@login_required
def repadminfinan(request):
    materiapr = tablarepmateriapr(request)
    compramed = tablarepcompramed(request) 
    preciocanal = tablareppreciocanal(request) 
    nuevosclientes = tablarepnuevosclientes(request) 
    return render(request, 'report_gadminfinan.html', {'materiapr': materiapr,'compramed':compramed,'preciocanal':preciocanal,'nuevosclientes':nuevosclientes})

def tablarepmateriapr(request):
    with connections['b_gaf'].cursor() as cursor:
        cursor.execute('''SELECT MATERIA_PRIMA,COSTO_PROMEDIO,CANTIDAD_COMPRADA,DIAS_INVENTARIO,FECHA_CORTE FROM b_gaf.compras_materia_prima''')
        materiapr = cursor.fetchall()   
    return materiapr
def tablarepcompramed(request):
    with connections['b_gaf'].cursor() as cursor:
        cursor.execute('''SELECT VALOR,MEDICAMENTO,CLASIFICACION,CANTIDAD,TIPO,FECHA_CORTE FROM b_gaf.compras_medicamentos''')
        compramed = cursor.fetchall()   
    return compramed
def tablareppreciocanal(request):
    with connections['b_gaf'].cursor() as cursor:
        cursor.execute('''SELECT NIT,CLIENTE,ZONA,VALOR,SEMANA  FROM b_gaf.precio_canales_semana WHERE GUID = (SELECT MAX(GUID) FROM b_gaf.precio_canales_semana)''')
        preciocanal = cursor.fetchall()   
    return preciocanal
def tablarepnuevosclientes(request):
    with connections['dhc'].cursor() as cursor:
        cursor.execute('''SELECT NIT,RAZON_SOCIAL,CUPO,DIRECCION_SEDE_PRINCIPAL,DIRECCION_EXPENDIO,ID_CLASIFICACION,ID_MUNICIPIO,ID_DEPARTAMENTO,ID_REGION,ID_VENDEDOR,ID_SEGMENTO,ID_linea_negocio FROM dhc.clientes;''')
        nuevosclientes = cursor.fetchall()   
    return nuevosclientes

#---------------- TABLAS DE REPORTES GESTION HUMANA------------------------------------------
#---------------------------------------------------------------------------------------------
@never_cache
@login_required
def repgestionhumana(request):
    nomina = tablarepnomina(request)
    promociones = tablareppromociones(request)
    procesosele = tablarepprocesosele(request) 
    retencion = tablarepretencion(request) 
    rotacion = tablareprotacion(request) 
    sstdiagindi = tablarepsstdiagindi(request) 
    sstindi = tablarepsstindi(request) 
    sstseveridad = tablarepsstseveridad(request)
    recupnomina = tablareprecunomina(request)
    return render(request, 'report_gestionhumana.html', {'nomina': nomina,'promociones': promociones,'procesosele':procesosele,'retencion':retencion,'rotacion':rotacion,'sstdiagindi':sstdiagindi,'sstindi':sstindi,'sstseveridad':sstseveridad,'recupnomina':recupnomina})

def tablarepnomina(request):
    with connections['b_gh'].cursor() as cursor:
        cursor.execute('''SELECT FECHA_CORTE,AREA,CENTRO_COSTO,NUM_COLABORADORES,COSTO_PROV FROM b_gh.nomina''')
        nomina = cursor.fetchall()
      
    return nomina
def tablareprecunomina(request):
    with connections['b_gh'].cursor() as cursor:
        cursor.execute('''SELECT VALOR_RECUPERADO,FECHA_CORTE FROM b_gh.recuperado_nomina WHERE GUID = (SELECT MAX(GUID) FROM b_gh.recuperado_nomina)''')
        recupnomina = cursor.fetchall()
      
    return recupnomina
def tablareppromociones(request):
    with connections['b_gh'].cursor() as cursor:
        cursor.execute('''SELECT FECHA_CORTE,NOMBRE,ANTIGUO_CARGO,NUEVO_CARGO FROM b_gh.promociones''')
        promociones = cursor.fetchall()
    return promociones
def tablarepprocesosele(request):
    with connections['b_gh'].cursor() as cursor:
        cursor.execute('''SELECT NUM_REQUISICION,FECHA_APROBACION,AREA_CENTRO_COSTO,FECHA_RETIRO,NOMBRE_RETIRADO,CARGO,CUBRIMIENTO_ESPERADO_DIAS,NOMBRE_CANDIDATO,TIPO_INGRESO_PROMO_INT,EXAMEN_MEDICO,VISITA_DOMICILIARIA,POLIGRAFIA,FECHA_INGRESO FROM b_gh.proceso_seleccion ''')
        procesosele = cursor.fetchall()   
    return procesosele
def tablarepretencion(request):
    with connections['b_gh'].cursor() as cursor:
        cursor.execute('''SELECT FECHA_REPORTE,INDICADOR_RETENCION,OBSERVACIONES FROM b_gh.retencion''')
        retencion = cursor.fetchall()   
    return retencion
def tablareprotacion(request):
    with connections['b_gh'].cursor() as cursor:
        cursor.execute('''SELECT FECHA_REPORTE,INDICADOR_ROTACION,OBSERVACIONES FROM b_gh.rotacion''')
        rotacion = cursor.fetchall()   
    return rotacion
def tablarepsstdiagindi(request):
    with connections['b_gh'].cursor() as cursor:
        cursor.execute('''SELECT FECHA_CORTE,SEDE,DIAGNOSTICO,CANTIDAD,OBSERVACION FROM b_gh.sst_diagnosticos_indicadores ''')
        sstdiagindi = cursor.fetchall()   
    return sstdiagindi
def tablarepsstindi(request):
    with connections['b_gh'].cursor() as cursor:
        cursor.execute('''SELECT FECHA_CORTE,SEDE,CANTIDAD_PEG,DIAS_INCAPACIDAD_PEL,CANTIDAD_PAT,PRORROGAS,DIAS_INCAPACIDAD_PAT,LICENCIA_MATERNIDAD,DIAS_LICENCIA_MAT,LICENCIA_PATERNIDAD,DIAS_LICENCIA_PAT,COSTO_INCAPACIDAD,OBSERVACIONES FROM b_gh.sst_indicadores''')
        sstindi = cursor.fetchall()   
    return sstindi
def tablarepsstseveridad(request):
    with connections['b_gh'].cursor() as cursor:
        cursor.execute('''SELECT FECHA_CORTE,CANT_ENF_GENERAL,CANT_ACC_TRABAJO,NUM_EMPLEADOS,FREC_ACC,DIAS_INC_GENERAL,DIAS_INC_ACC,SEV_ACC,INCID_ENF_LAB,PORC_AUSENTISMO FROM b_gh.sst_severidad_y_frecuencia''')
        sstseveridad = cursor.fetchall()   
    return sstseveridad


















# --------VISTAS PEDIDO GRANJA ----------------------------------------









from django.db.models import Q



@never_cache
@login_required
def repremision(request):
    consecutivo_cercafe = request.GET.get('consecutivoCercafe', None)
    if consecutivo_cercafe:
        remisionnew = tablaremisionnew(consecutivo_cercafe)
        print(consecutivo_cercafe)
        return JsonResponse({'remisionnew': remisionnew})
    else:
        # Si no se proporciona un consecutivo, simplemente renderiza la plantilla HTML
        return render(request, 'remision.html')

@never_cache
@login_required
def mortalidad(request):
    return render(request, 'mortalidad.html', {'granjas': granjas,'sitio': sitio})


@never_cache
@login_required
def disponiblilidad(request):
    return render(request, 'disponible.html', {'granjas': granjas})

@never_cache
@login_required
def pedido_granja(request):
    return render(request, 'pedido_granja.html', {'granjas': granjas})

def get_week_range(date):
    if isinstance(date, str):
        date = dt.strptime(date, '%Y-%m-%d')
    
    # Ajustar para semana domingo-sábado
    current_day = date.weekday()  # 0 = Lunes, 6 = Domingo
    if current_day == 6:  # Si es domingo
        start_of_week = date
    else:
        # Retroceder hasta el domingo anterior
        days_to_sunday = (current_day + 1) % 7
        start_of_week = date - timedelta(days=days_to_sunday)
    
    end_of_week = start_of_week + timedelta(days=6)  # Sábado
    
    return start_of_week, end_of_week
@login_required
def verificar_disponibilidad(request):
    fecha = request.GET.get('fecha')
    granja = request.GET.get('granja')
    
    start_of_week, end_of_week = get_week_range(fecha)
    
    with connections['prodsostenible'].cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(*) 
            FROM disponiblidad_semanal 
            WHERE granja = %s 
            AND fecha_disponibilidad BETWEEN %s AND %s
        """, [granja, start_of_week, end_of_week])
        
        count = cursor.fetchone()[0]
        
    return JsonResponse({'existe': count > 0})

@never_cache
@never_cache
@login_required
def guardar_disponibilidad(request):
    if request.method == 'POST':
        granja = request.POST.get('granja').upper()
        fecha_disponibilidad = request.POST.get('fecha_disponibilidad')
        
        # Verificar si ya existe una disponibilidad para esta granja en la misma semana
        start_of_week, end_of_week = get_week_range(fecha_disponibilidad)
        
        with connections['prodsostenible'].cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*) 
                FROM disponiblidad_semanal 
                WHERE granja = %s 
                AND fecha_disponibilidad BETWEEN %s AND %s
            """, [granja, start_of_week, end_of_week])
            
            count = cursor.fetchone()[0]
            
            if count > 0:
                return JsonResponse({'status': 'error', 'message': 'Ya existe una disponibilidad para esta granja en la semana seleccionada'})
        
            # Si no existe, continuar con el guardado
            caracteristica = request.POST.get('caracteristica').upper()
            genero = request.POST.get('genero').upper()
            disponibilidad_cantidad = request.POST.get('disponibilidad_cantidad').upper()
            disponibilidad_restante = disponibilidad_cantidad
            peso_promedio_limite_inferior = request.POST.get('peso_promedio_limite_inferior').upper()
            peso_promedio_limite_superior = request.POST.get('peso_promedio_limite_superior').upper()
            observaciones = request.POST.get('observaciones', '').upper()
            guid = str(uuid.uuid4())
            usuario = request.user.username

            try:
                with connections['prodsostenible'].cursor() as cursor:
                    cursor.execute("""
                        INSERT INTO disponiblidad_semanal 
                        (granja, fecha_disponibilidad, caracteristica, genero, disponibilidad_cantidad, 
                        disponibilidadRestante, peso_promedio_limite_inferior, peso_promedio_limite_superior, 
                        observaciones, GUID, USUARIO) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, [granja, fecha_disponibilidad, caracteristica, genero, disponibilidad_cantidad,
                        disponibilidad_restante, peso_promedio_limite_inferior, peso_promedio_limite_superior,
                        observaciones, guid, usuario])
                    
                return JsonResponse({'status': 'success', 'message': 'Disponibilidad subida exitosamente'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Error al guardar la disponibilidad: {str(e)}'})

    return render(request, 'disponible.html')


@never_cache
@login_required
def guardar_mortalidad(request):
    if request.method == 'POST':
            sitio = request.POST.get('sitio').upper()
            granja = request.POST.get('granja')
            tipo_salida = request.POST.get('tipo_salida')
            lote = request.POST.get('lote')
            fecha_salida = request.POST.get('fecha_salida')
            cantidad_cerdos = request.POST.get('cantidad_cerdos')
            peso = request.POST.get('peso')
            guid = str(uuid.uuid4())
            usuario = request.user.username

            try:
                with connections['prodsostenible'].cursor() as cursor:
                    cursor.execute("""
                        INSERT INTO mortalidad 
                        (sitio, granja, tipo_salida, lote, fecha_salida, cantidad_cerdos, peso, GUID, USUARIO) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, [sitio,granja,tipo_salida,lote, fecha_salida , cantidad_cerdos, peso, guid, usuario])

                return JsonResponse({'status': 'success', 'message': 'Mortalidad Guardada Exitosamente'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Error al guardar la mortalidad: {str(e)}'})

    return render(request, 'mortalidad.html')



@never_cache
@login_required
def disponibilidad_semanal(request):
    if request.method == "GET" and "FechaInicio" in request.GET and "FechaFin" in request.GET:
        fecha_inicio = request.GET.get('FechaInicio')
        fecha_fin = request.GET.get('FechaFin')

        # Convertir las fechas de texto a formato datetime
        fecha_inicio = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d')

        with connections['prodsostenible'].cursor() as cursor:
            cursor.execute("""
                SELECT 
                    g.ID, g.GRANJAS, ds.id, ds.fecha_disponibilidad, ds.disponibilidad_cantidad, 
                    di.id AS id_disponibilidad_individual, di.fechaDisponibilidad, di.cantidadCerdos
                FROM 
                    prodsostenible.disponiblidad_semanal ds
                JOIN 
                    dhc.granjas g ON ds.granja = g.ID
                LEFT JOIN 
                    prodsostenible.disponibilidadindividual di 
                ON 
                    di.consecutivoDisponibilidad = ds.id
                WHERE DATE 
                   (ds.fecha_disponibilidad) BETWEEN %s AND %s
            """, [fecha_inicio, fecha_fin])
            rows = cursor.fetchall()

            # Estructurar los datos para la tabla
            data = []
            for row in rows:
                data.append({
                    'granja_id': row[0],
                    'nombre_granja': row[1],  
                    'consecutivoDisponibilidad': row[2], 
                    'fecha_disponibilidad': row[3],  
                    'disponibilidad_cantidad': row[4],
                    'id_disponibilidad_individual': row[5],
                    'fechaDisponibilidad': row[6],
                    'cantidadCerdos': row[7],  # cantidad de cerdos pedidos
                })

            return JsonResponse(data, safe=False)

    return render(request, 'pedido_granja.html')


@never_cache
@login_required
@csrf_exempt
def solicitar_pedido(request):
    if request.method == 'POST':
        data = json.loads(request.POST.get('productos'))
        granja_id = request.POST.get('granja_id').upper()
        consecutivoDisponibilidad = request.POST.get('consecutivoDisponibilidad').upper()

        guid = str(uuid.uuid4())
        
        # Obtener el usuario autenticado
        usuario = request.user.username

        try:
            with connections['prodsostenible'].cursor() as cursor:
                for producto in data:
                    cantidadCerdos = int(producto['cantidad'])
                    frigorifico = producto['frigorifico'].upper()
                    fechaDisponibilidad = producto['fecha'].upper()
                    observacion = producto['observacion'].upper()

                    cursor.execute("INSERT INTO disponibilidadindividual (consecutivoDisponibilidad,granja_id, cantidadCerdos, frigorifico, fechaDisponibilidad, observacion, GUID, USUARIO) VALUES (%s,%s, %s, %s, %s, %s, %s, %s)", [consecutivoDisponibilidad,granja_id, cantidadCerdos, frigorifico, fechaDisponibilidad, observacion, guid, usuario])
                    

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Método no permitido'})



@never_cache
@login_required
def repdespacho(request):
    despachos = tabladespachos(request)
    # Obtener placas
    with connections['dhc'].cursor() as cursor:
        cursor.execute('SELECT id, placa FROM dhc.placas')
        placas = [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]
    
    with connections['dhc'].cursor() as cursor:
        cursor.execute('SELECT id, conductor FROM dhc.Conductores')
        conductores = [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]


    
    return render(request, 'despacho_frigos.html', {
        'despachos': despachos,
        'placas': placas,
        'conductores': conductores,
        
    })


@csrf_exempt
@login_required
def tabladespachos(request):
    try:
        with connections['prodsostenible'].cursor() as cursor:
            cursor.execute('''
                SELECT 
                    di.id,
                    di.consecutivoDisponibilidad, 
                    g.id AS granja_id, 
                    g.granjas AS granja_nombre, 
                    di.cantidadCerdos,
                    f.id AS frigorifico_id, 
                    f.nombre AS frigorifico_nombre, 
                    DATE_FORMAT(di.fechaDisponibilidad, '%d/%m/%Y') AS fechaDisponibilidad, 
                    di.observacion,
                    CAST((di.cantidadCerdos - IFNULL(SUM(dlg.cerdosDespachados), 0)) AS SIGNED) AS cerdos_sin_despachar
                FROM 
                    prodsostenible.disponibilidadindividual di
                LEFT JOIN 
                    dhc.granjas g ON di.Granja_id = g.id
                LEFT JOIN 
                    dhc.frigorificos f ON di.frigorifico = f.id
                LEFT JOIN 
                    despachoLotesGranjas dlg ON dlg.consecutivo_cercafe = di.id
                WHERE
                    (di.estado = 0 OR di.estado IS NULL)
                    AND di.fechaDisponibilidad >= CURDATE() - INTERVAL 15 DAY
                GROUP BY 
                    di.id, g.id, g.granjas, di.cantidadCerdos, f.nombre, di.fechaDisponibilidad, di.observacion
                ORDER BY 
                    di.id DESC
            ''')
            despachos = cursor.fetchall()
          
        return despachos
    except Exception as e:
        print(f"Error en la consulta: {e}")
        return []





@login_required
@csrf_exempt
@never_cache
def registrar_despacho(request):
    if request.method == 'POST':
        try:
            consecutivo = request.POST.get('consecutivoDisponibilidad')
            id = request.POST.get('consecutivoDespacho')
            granja_id = request.POST.get('granja_id')
            lote = request.POST.get('lote').upper()
            cerdos_despachados = int(request.POST.get('cerdosDespachados'))
            frigorifico = request.POST.get('frigorifico_id').upper()
            fecha_entrega = request.POST.get('fechaEntrega').upper()
            peso_total = request.POST.get('pesoTotal')
            placa = request.POST.get('placa').upper()
            regic = request.POST.get('regic', '')
            regica = request.POST.get('regica', '')
            retiro_alimento = request.POST.get('retiroalimento', '').upper()
            conductor = request.POST.get('conductor', '').upper()
            edad_prom = request.POST.get('edadprom', '')
            guid = str(uuid.uuid4())
            usuario = request.user.username
            fecha_actual = datetime.date.today()

            # Validar que la fecha de entrega no sea mayor a la fecha actual
            try:
                fecha_entrega_dt = datetime.datetime.strptime(fecha_entrega, '%Y-%m-%d').date()
                if fecha_entrega_dt > fecha_actual:
                    return JsonResponse({'success': False, 'error': 'La fecha de entrega no puede ser mayor a la fecha actual.'})
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Formato de fecha de entrega inválido.'})

            # Validar que la fecha de retiro de alimento no sea mayor a la fecha actual
            if retiro_alimento:
                try:
                    retiro_alimento_dt = datetime.datetime.strptime(retiro_alimento, '%Y-%m-%dT%H:%M').date()
                    if retiro_alimento_dt > fecha_actual:
                        return JsonResponse({'success': False, 'error': 'La fecha de retiro de alimento no puede ser mayor a la fecha actual.'})
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Formato de fecha de retiro de alimento inválido.'})



            # Verificar cuántos cerdos quedan por despachar para este pedido
            with connections['prodsostenible'].cursor() as cursor:
                cursor.execute('''
                    SELECT CAST((di.cantidadCerdos - IFNULL(SUM(dlg.cerdosDespachados), 0)) AS SIGNED) AS cerdos_sin_despachar
                    FROM prodsostenible.disponibilidadindividual di
                    LEFT JOIN despachoLotesGranjas dlg ON dlg.consecutivo_cercafe = di.id
                    WHERE di.id = %s
                    GROUP BY di.id
                ''', [id])
                cerdos_sin_despachar = cursor.fetchone()[0]

            with connections['prodsostenible'].cursor() as cursor:
                cursor.execute('''
                    INSERT INTO despachoLotesGranjas 
                    (ConsecutivoDisponibilidad, consecutivo_cercafe, granja, lote, cerdosDespachados, frigorifico, fechaEntrega, pesoTotal, placa, regic, regica, retiroalimento, conductor, edadprom,GUID, USUARIO) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s)
                ''', [
                    consecutivo, id, granja_id, lote, cerdos_despachados, frigorifico, fecha_entrega, peso_total, placa, regic, regica, retiro_alimento, conductor, edad_prom,guid,usuario
                ])

            actualizar_disponibilidad_restante(consecutivo, cerdos_despachados)

            return JsonResponse({'success': True, 'message': 'Despacho registrado correctamente.'})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

from django.db import connections, transaction
 
# Función para actualizar disponibilidadRestante
def actualizar_disponibilidad_restante(consecutivoDisponibilidad, cerdosDespachados):
    with connections['prodsostenible'].cursor() as cursor:
        cursor.execute("""
            SELECT disponibilidadRestante, disponibilidad_cantidad 
            FROM disponiblidad_semanal 
            WHERE id = %s
        """, [consecutivoDisponibilidad])
        disponibilidad = cursor.fetchone()
        
        if disponibilidad:
            # Explicitly convert to int to ensure type compatibility
            disponibilidadRestante = int(disponibilidad[0]) if disponibilidad[0] is not None else 0
            disponibilidad_cantidad = int(disponibilidad[1]) if disponibilidad[1] is not None else 0
            
            nueva_disponibilidad_restante = disponibilidadRestante - cerdosDespachados
             
            cursor.execute("""
                UPDATE disponiblidad_semanal
                SET disponibilidadRestante = %s
                WHERE id = %s
            """, [nueva_disponibilidad_restante, consecutivoDisponibilidad])
            transaction.commit()



@csrf_exempt
@login_required
@never_cache
def finalizar_registro(request):
    if request.method == 'POST':
        id = request.POST.get('id')  

        if not id:
            return JsonResponse({'success': False, 'error': 'No se recibió el Consecutivo Cercafe'})

        try:
            # Actualizar el estado en la base de datos
            with connections['prodsostenible'].cursor() as cursor:
                cursor.execute('''
                    UPDATE prodsostenible.disponibilidadindividual
                    SET estado = 1
                    WHERE id = %s
                ''', [id])

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Método no permitido'})



@never_cache
@login_required
def get_pedido(request):
    consecutivo_id = request.GET.get('consecutivoDisponibilidad')
    try:
        with connections['prodsostenible'].cursor() as cursor:
            cursor.execute('''
                SELECT id, cantidadCerdos, frigorifico, fechaDisponibilidad, observacion 
                FROM disponibilidadindividual
                WHERE consecutivoDisponibilidad = %s
            ''', [consecutivo_id])
            rows = cursor.fetchall()

            # Estructura los datos para enviar en la respuesta
            data = []
            for row in rows:
                data.append({
                    'id': row[0],
                    'cantidadCerdos': row[1],
                    'frigorifico': row[2],
                    'fechaDisponibilidad': row[3],
                    'observacion': row[4]
                })

            return JsonResponse({'pedidos': data}) # Devuelve la respuesta en formato JSON
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500) # Maneja correctamente los errores





@never_cache
@csrf_exempt
@login_required
def update_pedido(request):
    if request.method == 'POST':
        pedidos = json.loads(request.POST.get('pedidos'))  # Recibe los datos de los pedidos

        try:
            with connections['prodsostenible'].cursor() as cursor:
                for pedido in pedidos:
                   
                    frigorifico_id = int(pedido['frigorifico'])
                    
                    cursor.execute('''
                        UPDATE disponibilidadindividual
                        SET cantidadCerdos = %s, frigorifico = %s, fechaDisponibilidad = %s, observacion = %s
                        WHERE id = %s
                    ''', [
                        pedido['cantidad'],  # Cantidad de cerdos
                        frigorifico_id,      # ID del frigorífico
                        pedido['fecha'],     # Fecha de disponibilidad
                        pedido['observacion'],  # Observaciones
                        pedido['id']         # ID del pedido
                    ])

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)









#----- generar tabla de remisiones solo visualizacion de  informacion
def tablaremisionnew(consecutivo_cercafe):
    intranetcercafe2_connection = connections['prodsostenible']
    with intranetcercafe2_connection.cursor() as cursor:
        if consecutivo_cercafe:
            cursor.execute("SELECT ConsecutivoDisponibilidad,consecutivo_cercafe,granja,lote,cerdosDespachados,frigorifico,fechaEntrega,pesoTotal,conductor,placa,regic,regica,retiroalimento,edadprom from prodsostenible.despachoLotesGranjas WHERE consecutivo_cercafe = %s", [consecutivo_cercafe])
        else:
            cursor.execute("SELECT ConsecutivoDisponibilidad,consecutivo_cercafe,granja,lote,cerdosDespachados,frigorifico,fechaEntrega,pesoTotal,conductor,placa,regic,regica,retiroalimento,edadprom from prodsostenible.despachoLotesGranjas")
        remisionnew = cursor.fetchall()
    return remisionnew

#--- filtar por fechas para descargar el excel
def filtered_data(start_date, end_date):
    with connections['prodsostenible'].cursor() as cursor:
        cursor.execute('''
            SELECT ConsecutivoDisponibilidad, consecutivo_cercafe, granja, lote, cerdosDespachados, frigorifico, 
            fechaEntrega, pesoTotal, conductor, placa, regic, regica, retiroalimento, edadprom 
            FROM prodsostenible.despachoLotesGranjas
            WHERE fechaentrega BETWEEN %s AND %s
        ''', [start_date, end_date])
        fechas = cursor.fetchall()

    
    logger.info(fechas)

    return fechas


#-- generar excel de remisiones para reporte
def generar_excel(request):
    # Obtener las fechas de inicio y fin del request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Obtener los datos filtrados utilizando la función de la segunda vista
    compromisos = filtered_data(start_date, end_date)

    # Obtener el directorio de descargas del usuario actual
    downloads_dir = os.path.join(settings.BASE_DIR, 'tmp')

    # Crear el archivo Excel en el directorio de descargas
    filename = 'remisiones.xlsx'
    file_path = os.path.join(downloads_dir, filename)

    # Escribir los datos en el archivo Excel
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()

    # Escribir los encabezados
    headers = ['ConsecutivoDisponibilidad', 'consecutivo_cercafe', 'granja', 'lote', 'cerdosDespachados', 'frigorifico', 
               'fechaEntrega', 'pesoTotal', 'conductor', 'placa', 'regic', 'regica', 'retiroalimento', 'edadprom']
    for i, header in enumerate(headers):
        worksheet.write(0, i, header)

    # Obtener el índice de la columna 'fechaEntrega'
    fecha_entrega_index = headers.index('fechaEntrega') if 'fechaEntrega' in headers else None

    # Escribir los datos
    for row, compromiso in enumerate(compromisos, start=1):
        for col, value in enumerate(compromiso):
            # Formatear la fecha como un string solo si es un objeto datetime
            if fecha_entrega_index is not None and col == fecha_entrega_index:
                if isinstance(value, datetime.datetime):
                    value = value.strftime('%Y-%m-%d')
                # Si value es un str, no se modifica
            worksheet.write(row, col, value)

    # Cerrar el archivo Excel
    workbook.close()

    # Enviar el archivo Excel como respuesta de descarga
    with open(file_path, 'rb') as file:
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


#------ VISTA para creacion del PDF EN  REMISIONES------------
from django.template.loader import render_to_string
from django.http import HttpResponse
import pdfkit
import qrcode
from django.db import connections
from django.template.loader import render_to_string

def generate_qr_code(input_data):
    # Obtener la ruta absoluta del directorio 'static/images' dentro de Django
    static_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'static'))
    images_dir = os.path.join(static_dir, 'images')
    filename = 'qrcercafe.png'
    filepath = os.path.join(images_dir, filename)

    # Crea el directorio si no existe
    os.makedirs(images_dir, exist_ok=True)

    # Crea el código QR
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(input_data)
    qr.make(fit=True)

    # Crea la imagen del código QR
    img = qr.make_image(fill='black', back_color='white')

    # Guardar la imagen en la ruta especificada
    img.save(filepath)
    
    
def generar_pdf(request):
    # Conexiones de bases de datos
    intranetcercafe2_connection = connections['prodsostenible']
    dhc_connection = connections['dhc'] 
    consecutivo_cercafe = request.GET.get('consecutivoCercafe', None)
    
    if consecutivo_cercafe: 
        # Obtener los datos de la remisión filtrados por el consecutivo ceracafe
        remisiones = tablaremisionnew(consecutivo_cercafe)
        
        # Consulta en la tabla despachoLotesGranjas
        with intranetcercafe2_connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    d.ConsecutivoDisponibilidad, 
                    d.consecutivo_cercafe, 
                    d.granja, 
                    d.lote, 
                    d.cerdosDespachados, 
                    d.frigorifico, 
                    d.fechaEntrega, 
                    d.pesoTotal, 
                    c.Conductor AS ConductorID, 
                    p.placa AS PlacaID, 
                    d.regic, 
                    d.regica, 
                    d.retiroalimento, 
                    d.edadprom 
                FROM 
                    despachoLotesGranjas d
                LEFT JOIN 
                    dhc.Conductores c ON d.conductor = c.id
                LEFT JOIN 
                    dhc.placas p ON d.placa = p.ID
                WHERE 
                    d.consecutivo_cercafe = %s
            """, [consecutivo_cercafe])
            remisionnew = cursor.fetchall()
        
        # Verificar si hay datos y seleccionar los últimos
        if remisionnew:
            ultimo_registro = remisionnew[-1]
            granja_primera_consulta = ultimo_registro[2]
            conductor_nombre = ultimo_registro[8]
            placa_nombre = ultimo_registro[9]
            regic = ultimo_registro[10]
            regica = ultimo_registro[11]
            retiro_alimento = ultimo_registro[12]
        
            # Consulta adicional en dhc.granjas
            with dhc_connection.cursor() as cursor:
                cursor.execute("""
                    SELECT
                        C.ID AS ID,
                        C.ID AS ID_INTRANET,
                        UPPER(C.GRANJAS) AS Granja,
                        D.CODIGO AS Nit_asociado,
                        UPPER(E.razon_social) AS Asociado
                    FROM
                        dhc.granjas C
                    JOIN 
                        dhc.nombre_comercial D ON C.NOMBRE_COMERCIAL = D.ID
                    JOIN 
                        dhc.razon_social E ON C.RAZON_SOCIAL = E.ID
                    WHERE 
                        UPPER(C.ID) = %s;
                """, [granja_primera_consulta])
                resultados_dhc = cursor.fetchall()
            
            frigorifico_id = ultimo_registro[5]
            nombre_frigorifico = None
            
            if frigorifico_id:
                with dhc_connection.cursor() as cursor:
                    cursor.execute("SELECT nombre FROM frigorificos WHERE ID = %s", [frigorifico_id])
                    nombre_frigorifico_result = cursor.fetchone()
                    if nombre_frigorifico_result:
                        nombre_frigorifico = nombre_frigorifico_result[0]
            




            
            total_cantidad = sum(remision[7] for remision in remisionnew)
            total_cerdos = sum(remision[4] for remision in remisionnew)
            promedio = total_cantidad / total_cerdos if total_cerdos > 0 else 0
            promedio_formateado = f'{promedio:.2f}'
            
            input_data = (resultados_dhc[0][0], resultados_dhc[0][2], consecutivo_cercafe, str(total_cerdos), str(total_cantidad), placa_nombre, regica, resultados_dhc[0][3])
            generate_qr_code(input_data)
            
            html = render_to_string('remision_pdf.html', {
                'remisiones': remisiones,
                'promedio_formateado': promedio_formateado,
                'remisionnew': remisionnew,
                'resultados_dhc': resultados_dhc,
                'consecutivo_cercafe': consecutivo_cercafe,
                'total_cantidad': total_cantidad,
                'total_cerdos': total_cerdos,
                'nombre_frigorifico': nombre_frigorifico,
                'retiro_alimento': retiro_alimento,
                'conductor_nombre': conductor_nombre,
                'placa_nombre': placa_nombre,
            })
            
            pdf = pdfkit.from_string(html, False, configuration=pdfkit.configuration(wkhtmltopdf='C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe'))
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="reporte_remisiones.pdf"'
            return response
        
    return HttpResponse("No se proporcionó un consecutivo válido.")















































































































#*-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-
#*-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-
#*-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-
#*-*-*-*-*-*-*-*-*-*-*-*                    *-*-*-*-*-*-*-*-*-*-*-**-*-*-
#*-*-*-*-*-*-*-*-*-*-*-*        APIS        **-*-*-*-*-*-*-*-*-*-*-*-**-**-*
#*-*-*-*-*-*-*-*-*-*-*-*                    -**-*-*-*-*-*-*-*-*-*-*-**-*-*-
#*-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-
#*-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-
#*-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-*-*-*-*-*-*-*-*-**-*-*-


def api_hembras_registradas(request):
    # Obtener la conexión a la base de datos intranetcercafe2
    intranetcercafe2_connection = connections['int']

    # Realizar la consulta a la base de datos
    with intranetcercafe2_connection.cursor() as cursor:
        cursor.execute("SELECT * FROM hembras_registradas")
        results = cursor.fetchall()

    # Procesar los resultados y construir la respuesta JSON
    items = {'Ingreso_lote': []}
    for granja in results:
        item = {
            'id': granja[0],
            'id_lote': granja[1],
            'nombre_hembra': granja[2],
            'estado': granja[3]
        }
        items['Ingreso_lote'].append(item)

    response = {
        'success': True,
        'data': items,
        'message': 'data_hembras_peso_esperado'
    }

    return JsonResponse(response, status=200)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from django.db import connections



import requests
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

def obtener_token():
    url = "https://cercafe.yeminus.com/apicercafe/token"
    payload = {
        "userName": "API",
        "password": "4p1C3rcaf3",
        "grant_type": "password"
    }
    headers = {
        "Content-Type": "application/x-www-form-urlencoded"
    }

    response = requests.post(url, data=payload, headers=headers)
    if response.status_code == 200:
        return response.json().get("access_token")
    else:
        raise Exception("Error al obtener el token: " + response.text)

def obtener_informeinventario(token, fecha_corte):
    url = "https://cercafe.yeminus.com/apicercafe/app/informes/informespersonalizados/ejecutarconsultainforme"
    payload = {
        "crearArchivo": False,
        "idInforme": 27,
        "filtroVariableInforme": [],
        "variablesFiltro": [
            {
                "codigo": "FECHA_CORTE",
                "valor": fecha_corte
            }
        ]
    }
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "id_empresa": "01",  
        "usuario": "API"    
    }

    response = requests.post(url, json=payload, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception("Error al obtener el informe: " + response.text)

@csrf_exempt
def informe_view(request):
    if request.method == "POST" or request.method == "GET":
        try:
            token = obtener_token()
            fecha_corte = request.GET.get("fecha_corte", "01-10-2024")  
            informe = obtener_informeinventario(token, fecha_corte)
            return JsonResponse(informe)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    else:
        return JsonResponse({"error": "Invalid request method."}, status=400)

from django.shortcuts import render
from django.http import HttpResponse
import requests
import pandas as pd
from .forms import DateRangeForm

def get_decomisos():
    response = requests.get('https://intranet.cercafe.com/FRIGOTUN/public/apiDecomisos')
    if response.status_code == 200:
        return response.json()['data']['decomisos']
    else:
        return []
    pass

def filter_decomisos_by_date(decomisos, start_date, end_date):
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)
    return [decomiso for decomiso in decomisos if start_date <= pd.to_datetime(decomiso['fecha_decomiso']) <= end_date]
    pass

def to_excel(decomisos):
    df = pd.DataFrame(decomisos)
    filename = 'decomisos.xlsx'
    file_path = os.path.join(settings.MEDIA_ROOT, filename)
    df.to_excel(file_path, index=False)
    return file_path

def decomisos_view(request):
    if request.method == 'POST':
        form = DateRangeForm(request.POST)
        if form.is_valid():
            # Corrige aquí los nombres para acceder correctamente a la data del formulario
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']  # Corregido de 'end_post' a 'end_date'
            decomisos = get_decomisos()
            filtered_decomisos = filter_decomisos_by_date(decomisos, start_date, end_date)
            excel_file_path = to_excel(filtered_decomisos)
            
            if os.path.exists(excel_file_path):
                response = FileResponse(open(excel_file_path, 'rb'), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename="decomisos.xlsx"'
                return response
            else:
                return HttpResponse("Error al generar el archivo Excel", status=500)
    else:
        form = DateRangeForm()
    
    return render(request, 'decomisos.html', {'form': form})



from django.http import JsonResponse
from django.db import connections

def api_recepcion(request):
    # Conexión a la base de datos
    intranetcercafe2_connection = connections['prod_carnica']

    try:
        with intranetcercafe2_connection.cursor() as cursor:
            # Ejecutar la consulta para obtener todos los datos de la tabla recepcion
            cursor.execute("SELECT * FROM recepcion")
            results = cursor.fetchall()

            # Obtener los nombres de las columnas
            column_names = [col[0] for col in cursor.description]

        # Construir la respuesta JSON
        items = {'recepcion': []}
        for row in results:
            # Combinar columnas y valores para formar un diccionario para cada fila
            item = dict(zip(column_names, row))
            items['recepcion'].append(item)

        response = {
            'success': True,
            'data': items,
            'message': 'data_recepcion_completa'
        }
        return JsonResponse(response, status=200)
    except Exception as e:
        # Manejo de errores
        response = {
            'success': False,
            'message': f'Error retrieving data: {str(e)}'
        }
        return JsonResponse(response, status=500)

def api_proveeduria(request):
    # Conexión a la base de datos
    prodsostenible_connection = connections['prodsostenible']

    try:
        with prodsostenible_connection.cursor() as cursor:
            # Consulta para unir las tablas y filtrar por el año 2025
            query = """
            SELECT 
                dl.consecutivo_cercafe, 
                dl.granja, 
                dl.cerdosDespachados, 
                dl.lote, 
                dl.pesoTotal, 
                dl.FechaDatos
            FROM despachoLotesGranjas dl
            WHERE YEAR(dl.FechaDatos) = 2025
            """

            # Ejecutar la consulta
            cursor.execute(query)
            results = cursor.fetchall()

            # Obtener los nombres de las columnas
            column_names = [col[0] for col in cursor.description]

        # Construir la respuesta JSON
        items = {'proveeduria': []}
        for row in results:
            # Combinar columnas y valores para formar un diccionario para cada fila
            item = dict(zip(column_names, row))
            items['proveeduria'].append(item)

        response = {
            'success': True,
            'data': items,
            'message': 'data_proveeduria_completa'
        }
        return JsonResponse(response, status=200)
    except Exception as e:
        # Manejo de errores
        response = {
            'success': False,
            'message': f'Error retrieving data: {str(e)}'
        }
        return JsonResponse(response, status=500)

def api_beneficio(request):
    # Conexión a la base de datos
    intranetcercafe2_connection = connections['prod_carnica']

    try:
        with intranetcercafe2_connection.cursor() as cursor:
            # Ejecutar la consulta para obtener todos los datos de la tabla recepcion
            cursor.execute("SELECT * FROM beneficio")
            results = cursor.fetchall()

            # Obtener los nombres de las columnas
            column_names = [col[0] for col in cursor.description]

        # Construir la respuesta JSON
        items = {'recepcion': []}
        for row in results:
            # Combinar columnas y valores para formar un diccionario para cada fila
            item = dict(zip(column_names, row))
            items['recepcion'].append(item)

        response = {
            'success': True,
            'data': items,
            'message': 'data_beneficio_completa'
        }
        return JsonResponse(response, status=200)
    except Exception as e:
        # Manejo de errores
        response = {
            'success': False,
            'message': f'Error retrieving data: {str(e)}'
        }
        return JsonResponse(response, status=500)




def api_beneficio_auditoria(request):
    # Conexión a la base de datos
    intranetcercafe2_connection = connections['prod_carnica']

    try:
        with intranetcercafe2_connection.cursor() as cursor:
            # Ejecutar la consulta para obtener todos los datos de la tabla recepcion
            cursor.execute("SELECT * FROM auditoria_beneficio")
            results = cursor.fetchall()

            # Obtener los nombres de las columnas
            column_names = [col[0] for col in cursor.description]

        # Construir la respuesta JSON
        items = {'recepcion': []}
        for row in results:
            # Combinar columnas y valores para formar un diccionario para cada fila
            item = dict(zip(column_names, row))
            items['recepcion'].append(item)

        response = {
            'success': True,
            'data': items,
            'message': 'data_beneficio_completa'
        }
        return JsonResponse(response, status=200)
    except Exception as e:
        # Manejo de errores
        response = {
            'success': False,
            'message': f'Error retrieving data: {str(e)}'
        }
        return JsonResponse(response, status=500)

##########################    AGRINESS    ####################################################



class KpisReproductivosSitio1API(APIView):
    # Clases de autenticación: Aquí indicamos que usaremos TokenAuthentication
    authentication_classes = [TokenAuthentication]
    # Clases de permisos: Aquí indicamos que solo usuarios autenticados pueden acceder
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            # Conexión a la base de datos 'agriness'
            # Asegúrate de que 'agriness' es el nombre de la conexión en settings.py
            agriness_connection = connections['agriness']

            with agriness_connection.cursor() as cursor:
                # Ejecutar la consulta para obtener todos los datos de la tabla kpis_reproductivos_sitio1
                # Si 'agriness' es un esquema dentro de la DB, la consulta es correcta.
                # Si 'agriness' es el nombre de la DB y la tabla está en el esquema por defecto,
                # podrías solo usar "SELECT * FROM kpis_reproductivos_sitio1"
                cursor.execute("SELECT * FROM agriness.kpis_reproductivos_sitio1")
                results = cursor.fetchall()

                # Obtener los nombres de las columnas
                column_names = [col[0] for col in cursor.description]

                # Construir la respuesta JSON
                items = []
                for row in results:
                    item = dict(zip(column_names, row))
                    items.append(item)

            response_data = {
                'success': True,
                'data': items,
                'message': 'Datos de KPIs Reproductivos del sitio 1 obtenidos exitosamente.'
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            error_response = {
                'success': False,
                'message': f'Error al obtener datos de KPIs Reproductivos del sitio 1: {str(e)}'
            }
            return Response(error_response, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



    ######### DESPACHO DETALLE #######################################


class despacho_detalleAPI(APIView):
    # Clases de autenticación: Aquí indicamos que usaremos TokenAuthentication
    authentication_classes = [TokenAuthentication]
    # Clases de permisos: Aquí indicamos que solo usuarios autenticados pueden acceder
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            
            agriness_connection = connections['prod_carnica']

            with agriness_connection.cursor() as cursor:
               
                cursor.execute("SELECT * FROM prod_carnica.despacho_detalle")
                results = cursor.fetchall()

                
                column_names = [col[0] for col in cursor.description]

                
                items = []
                for row in results:
                    item = dict(zip(column_names, row))
                    items.append(item)

            response_data = {
                'success': True,
                'data': items,
                'message': 'Datos de despacho detalle obtenidos exitosamente.'
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            error_response = {
                'success': False,
                'message': f'Error al obtener datos de despacho detalle: {str(e)}'
            }
            return Response(error_response, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
